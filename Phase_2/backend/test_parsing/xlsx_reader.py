import asyncio
import hashlib
import io
import json
import logging
import os
import re
import tempfile
import zipfile
import subprocess
from collections import Counter
from collections import defaultdict
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple, Union
from xml.etree import ElementTree as ET
from pdf2image import convert_from_path
import numpy as np
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import (
    AbsoluteAnchor,
    OneCellAnchor,
    TwoCellAnchor,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image
from sklearn.cluster import DBSCAN

from llm_utils import detect_sheet_layout_type, flatten_table_to_hierarchy, call_llm
from llm_provider import LLMProvider

log = logging.getLogger(__name__)


from dotenv import load_dotenv

dotenv_path = os.path.join(os.path.dirname(__file__), ".env")
load_dotenv(dotenv_path=dotenv_path, override=True)


# Markers consistent with docx_reader.py
IMAGE_PATH_START_MARKER = "[START_IMAGE_PATH]"
IMAGE_PATH_END_MARKER = "[END_IMAGE_PATH]"
TABLE_START_MARKER = "[START_TABLE_CONTENT]"
TABLE_END_MARKER = "[END_TABLE_CONTENT]"
SHAPE_START_MARKER = "[START_SHAPE_CONTENT]"
SHAPE_END_MARKER = "[END_SHAPE_CONTENT]"

# XML namespaces for xlsx
NS_SPREADSHEETML = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_DRAWING = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
NS_DRAWING_MAIN = "http://schemas.openxmlformats.org/drawingml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_CHART = "http://schemas.openxmlformats.org/drawingml/2006/chart"
NS_CTRLPROP = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

# EMU conversion
EMU_PER_PIXEL = 9525
# Tool schema
LAYOUT_DETECTION_TOOL = {
    "name": "classify_sheet_layout",
    "type": "function",
    "description": "Classify Excel sheet layout type as either table or zone layout",
    "parameters": {
        "type": "object",
        "properties": {
            "layout_type": {
                "type": "string",
                "enum": ["table", "zone"],
                "description": "The detected layout type: 'table' for standard table layout, 'zone' for zone-divided layout with repeating headers"
            },
            "confidence": {
                "type": "number",
                "minimum": 0,
                "maximum": 1,
                "description": "Confidence score for the classification (0-1)"
            },
            "reasoning": {
                "type": "string",
                "description": "Brief explanation of the classification decision"
            }
        },
        "required": ["layout_type", "confidence", "reasoning"]
    }
}

# Tool schema
TABLE_FLATTEN_TOOL = {
    "name": "flatten_table_hierarchy",
    "type": "function",
    "description": "Convert table data into hierarchical text structure with indentation",
    "parameters": {
        "type": "object",
        "properties": {
            "hierarchical_text": {
                "type": "string",
                "description": "The flattened hierarchical representation with proper indentation showing relationships between properties"
            },
            "primary_keys": {
                "type": "array",
                "items": {"type": "string"},
                "description": "List of column/property names used as primary keys to identify unique entries"
            },
            "entry_count": {
                "type": "integer",
                "description": "Number of top-level entries identified in the table"
            }
        },
        "required": ["hierarchical_text", "primary_keys", "entry_count"]
    }
}

# Header detection tool schema
HEADER_DETECTION_TOOL = {
    "name": "detect_table_headers",
    "type": "function",
    "description": "Detect which rows contain table headers (column names). Can identify multi-level headers.",
    "parameters": {
        "type": "object",
        "properties": {
            "header_row_indices": {
                "type": "array",
                "items": {"type": "integer"},
                "description": "Zero-based indices of rows that contain headers (e.g., [0] for single header, [0, 1] for multi-level headers)"
            },
            "explanation": {
                "type": "string",
                "description": "Explanation of why these rows were identified as headers"
            }
        },
        "required": ["header_row_indices", "explanation"]
    }
}

class XlsxParser:
    """Parser for XLSX files that extracts cells, images, charts, shapes and builds document tree."""

    def __init__(
        self,
        get_table_content: bool = True,
        get_image_content: bool = True,
        get_shape_content: bool = True,
        get_chart_content: bool = True,
        detect_sheet_layout_with_llm: bool = True,
        dbscan_eps: float = 50.0,
        dbscan_min_samples: int = 1,
        llm: Any = None,
        llm_model: str = "gpt-4.1-mini",
    ):
        """
        Initialize XLSX parser.

        Args:
            get_table_content: Whether to extract table content
            get_image_content: Whether to extract images
            get_shape_content: Whether to extract shapes
            get_chart_content: Whether to extract charts
            dbscan_eps: DBSCAN epsilon parameter for clustering nearby elements
            dbscan_min_samples: DBSCAN min_samples parameter
            llm: LLM client for layout detection (optional)
            llm_params_fn: Function that returns provider-specific params per model
            llm_model: Model name to use for LLM calls
        """
        self.get_table_content = get_table_content
        self.get_image_content = get_image_content
        self.get_shape_content = get_shape_content
        self.get_chart_content = get_chart_content
        self.detect_sheet_layout_with_llm = detect_sheet_layout_with_llm
        self.dbscan_eps = dbscan_eps
        self.dbscan_min_samples = dbscan_min_samples
        self.llm = llm
        self.llm_model = llm_model
        self._image_dir: Optional[str] = None
        self._saved_images: List[str] = []
        self._xlsx_path: Optional[str] = None
        self._workbook = None

        if self.detect_sheet_layout_with_llm and self.llm is None:
            self.llm = LLMProvider.create()
            
    @staticmethod
    def _md5(data: bytes) -> str:
        """Calculate MD5 hash of data."""
        return hashlib.md5(data).hexdigest()

    def _llm_params(self, model: str) -> dict[str, str]:
        if os.getenv("PREFER_PROVIDER", "openai") == "openai":
            return {
                "model": model,
            }
        return {
            "model": os.getenv("AWS_LLM_MODEL", "mistral.mistral-large-2407-v1:0"),
            "region": os.getenv("AWS_LLM_REGION", "us-west-2"),
        }

    def _detect_sheet_layout_type_for_table(
        self,
        cell_data: Dict[str, Any],
        table: Dict[str, Any]
    ) -> str:
        """
        Detect layout type (table vs zone) based on percentage of empty cells.

        If >50% of cells are empty, it's classified as "zone" layout.
        Otherwise, it's classified as "table" layout.

        Args:
            cell_data: Cell data dict with data, borders, etc.
            table: Single table dict with min_row, max_row, min_col, max_col, data

        Returns:
            "table" or "zone"
        """
        table_data = table["data"]
        if not table_data or len(table_data) == 0:
            return "table"

        total_cells = 0
        empty_cells = 0

        for row in table_data:
            for cell in row:
                total_cells += 1
                cell_str = str(cell).strip()
                if not cell_str or cell_str.lower() in ['none', 'nan', '']:
                    empty_cells += 1

        empty_percentage = (empty_cells / total_cells * 100) if total_cells > 0 else 0

        layout_type = "zone" if empty_percentage >= 50 else "table"

        print(f'[DEBUG] Layout detection: {empty_cells}/{total_cells} empty cells ({empty_percentage:.1f}%) -> {layout_type}')

        return layout_type

    def _divide_zone_layout_into_tables(
        self,
        cell_data: Dict[str, Any],
        table: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """
        Divide a zone layout table into sub-tables by detecting rectangular zones using borders.

        Algorithm:
        1. Loop through each cell in the table
        2. For each unvisited cell:
           - Expand right until finding a cell with right border
           - Expand down until finding a cell with down border
           - Create zone rectangle from top-left to bottom-right
           - Mark all cells in zone as visited
           - Extract content: cells in row joined by ' ', rows joined by '\n'

        Args:
            cell_data: Cell data dict with data, borders, etc.
            table: Single table dict with min_row, max_row, min_col, max_col, data

        Returns:
            List of sub-tables with text content
        """
        data = table["data"]
        if not data or len(data) == 0:
            return []

        borders = cell_data["borders"]
        min_row = table["min_row"]
        max_row = table["max_row"]
        min_col = table["min_col"]
        max_col = table["max_col"]

        num_rows = len(data)
        num_cols = len(data[0]) if data else 0

        visited = [[False] * num_cols for _ in range(num_rows)]
        sub_tables = []

        has_border = lambda row, col, side: self._has_border(
            borders, row, col, side, min_row, max_row, min_col, max_col
        )

        for row_idx in range(num_rows):
            for col_idx in range(num_cols):
                if visited[row_idx][col_idx]:
                    continue
                abs_row = min_row + row_idx
                abs_col = min_col + col_idx
                right_col_idx = col_idx
                for c in range(col_idx, num_cols):
                    abs_c = min_col + c
                    right_col_idx = c
                    if has_border(abs_row, abs_c, 'right'):
                        break
                down_row_idx = row_idx
                for r in range(row_idx, num_rows):
                    abs_r = min_row + r
                    down_row_idx = r
                    if has_border(abs_r, abs_col, 'bottom'):
                        break
                zone_rows = []
                for r in range(row_idx, down_row_idx + 1):
                    if r >= num_rows:
                        break
                    zone_row_content = []
                    for c in range(col_idx, right_col_idx + 1):
                        if c >= num_cols:
                            break
                        visited[r][c] = True
                        cell_content = str(data[r][c]).strip()
                        zone_row_content.append(cell_content)
                    zone_rows.append(' '.join(zone_row_content))

                zone_text = '\n'.join(zone_rows)

                if zone_text.strip():
                    sub_tables.append({
                        "text": zone_text,
                        "layout_type": "zone",
                        "min_row": min_row + row_idx,
                        "max_row": min_row + down_row_idx,
                        "min_col": min_col + col_idx,
                        "max_col": min_col + right_col_idx,
                        "zone_bounds": {
                            "row_start": row_idx,
                            "row_end": down_row_idx,
                            "col_start": col_idx,
                            "col_end": right_col_idx
                        }
                    })

        print(f'[DEBUG] Divided zone into {len(sub_tables)} sub-tables')
        return sub_tables

    def _flatten_table_to_hierarchy(
        self,
        formatted_table_data: str,
        layout_type: str
    ) -> Optional[str]:
        """
        Flatten table data into hierarchical structure using LLM.

        Args:
            formatted_table_data: Formatted content for table, markdown with nested table formatted
            layout_type: Either "table" or "zone"

        Returns:
            Hierarchical text representation or None if LLM not available
        """
        if not self.detect_sheet_layout_with_llm:
            log.warning("LLM not configured. Cannot flatten table to hierarchy.")
            return None

        try:
            response = asyncio.run(
                flatten_table_to_hierarchy(
                    self.llm,
                    self._llm_params,
                    model=self.llm_model,
                    table_data=formatted_table_data,
                    layout_type=layout_type,
                    tools=[TABLE_FLATTEN_TOOL],
                    tool_choice={"type": "function", "name": "flatten_table_hierarchy"},
                )
            )
            for out in response:
                if out.type == "function_call" and out.name == "flatten_table_hierarchy":
                    result = json.loads(out.arguments or "{}")
                    hierarchical_text = result.get("hierarchical_text", "")
                    primary_keys = result.get("primary_keys", [])
                    entry_count = result.get("entry_count", 0)


                    return hierarchical_text
            log.warning("No tool call in LLM response for table flattening.")
            return None

        except Exception as exc:
            log.warning("LLM table flattening failed: %s", exc)
            return None

    def _forward_fill_table(self, table_data: List[List[str]]) -> List[List[str]]:
        """
        Forward fill empty cells with content from the cell above in the same column.

        Common pattern in Excel where merged cells or repeated values are left blank,
        implying they continue from the row above.

        Args:
            table_data: 2D list of table data

        Returns:
            Table data with empty cells filled
        """
        if not table_data or len(table_data) < 2:
            return table_data

        # deep copy
        filled_data = [row[:] for row in table_data]

        num_cols = len(filled_data[0]) if filled_data else 0

        for col_idx in range(num_cols):
            for row_idx in range(1, len(filled_data)):
                current_val = filled_data[row_idx][col_idx].strip()

                if not current_val or current_val.lower() in ['none', 'nan']:
                    if row_idx > 0 and col_idx < len(filled_data[row_idx - 1]):
                        above_val = filled_data[row_idx - 1][col_idx].strip()
                        if above_val and above_val.lower() not in ['none', 'nan']:
                            filled_data[row_idx][col_idx] = above_val

        return filled_data

    def _remove_empty_rows(self, table: Dict[str, Any]) -> None:
        """
        Remove rows that contain only empty cells from table data.

        This modifies the table in-place by filtering table["data"].
        The table bounds (min_row, max_row, min_col, max_col) are kept unchanged
        to preserve the original sheet positions.

        Args:
            table: Table dict with "data" field (2D list)
        """
        if "data" not in table or not table["data"]:
            return

        original_row_count = len(table["data"])

        filtered_data = []
        for row in table["data"]:
            has_content = False
            for cell in row:
                cell_str = str(cell).strip()
                if cell_str and cell_str.lower() not in ['none', 'nan', '']:
                    has_content = True
                    break

            if has_content:
                filtered_data.append(row)

        table["data"] = filtered_data

        removed_count = original_row_count - len(filtered_data)
        if removed_count > 0:
            print(f'[DEBUG] Removed {removed_count} empty rows from table. Rows: {original_row_count} -> {len(filtered_data)}')

    def _verify_header_rows(self, table_data: List[List[str]], detected_headers: List[int]) -> List[int]:
        """
        Verify LLM-detected header rows using empty cell count pattern.

        Strategy:
        1. Fill forward the detected header rows
        2. Count empty cells in each row
        3. Find subset where empty cell count strictly decreases
        4. Subset must start from 0 and be consecutive

        Rationale: Multi-level headers typically have:
        - Top level (row 0): More merged cells → more empty cells
        - Lower levels: Fewer merged cells → fewer empty cells
        - Data rows: May have many empty cells (not strictly decreasing pattern)

        Args:
            table_data: 2D list of table data
            detected_headers: List of header row indices from LLM

        Returns:
            Verified list of header row indices (subset of detected_headers)
        """
        if not detected_headers or not table_data:
            return [0] if table_data else []
        header_rows = [table_data[i] for i in detected_headers if i < len(table_data)]
        if not header_rows:
            return [0]
        filled_header_rows = self._forward_fill_table(header_rows)
        empty_counts = []
        for row in filled_header_rows:
            empty_count = 0
            for cell in row:
                cell_str = str(cell).strip()
                if not cell_str or cell_str.lower() in ['none', 'nan', '']:
                    empty_count += 1
            empty_counts.append(empty_count)

        verified_indices = [0]
        for i in range(1, len(empty_counts)):
            if empty_counts[i] < empty_counts[i - 1]:
                verified_indices.append(i)
            else:
                break

        verified_headers = [detected_headers[i] for i in verified_indices if i < len(detected_headers)]

        if verified_headers != detected_headers:
            print(f'[DEBUG] Header verification: LLM detected {detected_headers}, verified {verified_headers}')
            print(f'[DEBUG] Empty cell counts after fill-forward: {empty_counts}')

        return verified_headers

    def _detect_table_header_rows(self, table_data: List[List[str]], max_rows: int = 10) -> List[int]:
        """
        Detect which rows contain table headers using LLM.

        Args:
            table_data: 2D list of table data
            max_rows: Maximum number of rows to analyze (default: 10)

        Returns:
            List of zero-based consecutive indices of header rows (e.g., [0] or [0, 1] for multi-level headers)
        """
        if not self.detect_sheet_layout_with_llm or not self.llm:
            return [0] if table_data and len(table_data) > 0 else []

        sample_rows = table_data[:max_rows] if len(table_data) >= max_rows else table_data

        rows_text = ""
        for row_idx, row in enumerate(sample_rows):
            rows_text += f"Row {row_idx}: {row}\n"

        try:
            system_prompt = """You are an expert at analyzing table structures and identifying header rows.

IMPORTANT: Header rows are ALWAYS CONSECUTIVE starting from the top of the table.

Header rows contain column names/labels/descriptions that define what each column represents.
Data rows contain the actual values/content that correspond to those column definitions.

Header rows can be:
- Single row: One row containing all column names
- Multiple consecutive rows: Multiple rows that together form the complete column structure (e.g., grouped headers, hierarchical categories)

Key characteristics to identify headers vs data:
- Headers typically contain descriptive/categorical text that labels columns
- Data rows contain specific values, measurements, or entries
- In multi-level headers, upper rows contain broader categories, lower rows contain specific field names
- Headers are always at the top and consecutive (no data rows between header rows)

Your task: Identify which consecutive rows from the top are header rows (0-indexed)."""

            user_prompt = f"""Analyze the following table rows and identify which consecutive rows from the top are headers:

{rows_text}

CRITICAL: Header rows are consecutive and always start from row 0. Examples of valid header row patterns:
- Single header: [0]
- Two-level header: [0, 1]
- Three-level header: [0, 1, 2]

Invalid patterns (non-consecutive or not starting from 0): [1], [0, 2], [1, 2]

Analyze:
1. Which consecutive rows from the top contain column headers/labels?
2. Where does the actual data content begin?
3. Provide clear reasoning for your decision."""

            response = asyncio.run(
                call_llm(
                    self.llm,
                    self._llm_params,
                    model=self.llm_model,
                    system_prompt=system_prompt,
                    user_prompt=user_prompt,
                    tools=[HEADER_DETECTION_TOOL],
                    tool_choice={"type": "function", "name": "detect_table_headers"},
                    temperature=0,
                )
            )

            for out in response or []:
                if out.type == "function_call" and out.name == "detect_table_headers":
                    try:
                        result = json.loads(out.arguments or "{}")
                        header_indices = result.get("header_row_indices", [0])
                        explanation = result.get("explanation", "")
                        print(f'[DEBUG] LLM detected header rows: {header_indices} - {explanation}')
                        verified_indices = self._verify_header_rows(table_data, header_indices)
                        return verified_indices if verified_indices else [0]
                    except json.JSONDecodeError as e:
                        log.warning(f"Failed to parse header detection response: {e}")
                        continue
        except Exception as e:
            log.warning(f"LLM header detection failed: {e}. Defaulting to first row as header.")

        return [0]

    def _ensure_image_dir(self) -> str:
        """Create and return temp directory for images."""
        if self._image_dir is None:
            self._image_dir = tempfile.mkdtemp(prefix="xlsx_images_")
        return self._image_dir

    def _save_image_bytes(
        self, image_bytes: bytes, ext: Optional[str] = None, name_hint: str = ""
    ) -> Optional[Tuple[str, str]]:
        """
        Save image bytes to temp file.

        Returns:
            Tuple of (path, md5_hash) or None if failed
        """
        if not image_bytes:
            return None
        directory = self._ensure_image_dir()
        extension = (ext or "png").lstrip(".") or "png"
        image_hash = self._md5(image_bytes)
        filename = f"img_{len(self._saved_images) + 1}_{image_hash[:8]}.{extension}"
        path = os.path.join(directory, filename)
        try:
            Path(path).write_bytes(image_bytes)
        except Exception as exc:
            log.warning("Unable to write image %s: %s", name_hint, exc)
            return None
        self._saved_images.append(path)
        return path, image_hash

    def _emu_to_pixels(self, emu: int) -> float:
        """Convert EMU to pixels."""
        return emu / EMU_PER_PIXEL

    def _get_cell_pixel_position(
        self, ws: Worksheet, col: int, row: int
    ) -> Tuple[float, float]:
        """
        Get pixel position of a cell's top-left corner.

        Args:
            ws: Worksheet
            col: Column number (1-based)
            row: Row number (1-based)

        Returns:
            Tuple of (x, y) in pixels
        """
        x = 0.0
        for c in range(1, col):
            col_letter = get_column_letter(c)
            width = ws.column_dimensions[col_letter].width or 8.43  # xls internal default column width in pts
            x += width * 7 # pts to pixels

        y = 0.0
        for r in range(1, row):
            height = ws.row_dimensions[r].height or 15  # xls internal default column height in pts
            y += height * 1.33  # pts to pixels

        return x, y

    def _get_cell_size(
        self, ws: Worksheet, col: int, row: int
    ) -> Tuple[float, float]:
        """Get cell size in pixels."""
        col_letter = get_column_letter(col)
        width = ws.column_dimensions[col_letter].width or 8.43
        height = ws.row_dimensions[row].height or 15
        return width * 7, height * 1.33



    def _pixel_to_cell(
        self, ws: Worksheet, px_x: float, px_y: float
    ) -> Tuple[int, int]:
        """
        Convert pixel position to cell (row, col).

        Args:
            ws: Worksheet
            px_x: X position in pixels
            px_y: Y position in pixels

        Returns:
            Tuple of (row, col) - 1-based indices
        """
        col = 1
        accumulated_x = 0.0
        max_col = ws.max_column or 100
        for c in range(1, max_col + 1):
            col_letter = get_column_letter(c)
            col_width = (ws.column_dimensions[col_letter].width or 8.43) * 7
            if accumulated_x + col_width > px_x:
                col = c
                break
            accumulated_x += col_width
            col = c

        row = 1
        accumulated_y = 0.0
        max_row = ws.max_row or 100
        for r in range(1, max_row + 1):
            row_height = (ws.row_dimensions[r].height or 15) * 1.33
            if accumulated_y + row_height > px_y:
                row = r
                break
            accumulated_y += row_height
            row = r

        return row, col

    def _get_element_center_cell(
        self, ws: Worksheet, element: Dict[str, Any]
    ) -> Tuple[int, int]:
        """
        Get the cell where the center of an element is located.

        Args:
            ws: Worksheet
            element: Element dict with bbox or x, y, width, height

        Returns:
            Tuple of (row, col) - 1-based indices
        """
        if "bbox" in element:
            bbox = element["bbox"]
            center_x = (bbox[0] + bbox[2]) / 2
            center_y = (bbox[1] + bbox[3]) / 2
        else:
            x = element.get("x", 0)
            y = element.get("y", 0)
            width = element.get("width", 0)
            height = element.get("height", 0)
            center_x = x + width / 2
            center_y = y + height / 2

        return self._pixel_to_cell(ws, center_x, center_y)

    def _extract_checkboxes_from_xml(
        self, xlsx_path: str, sheet_name: str
    ) -> Dict[str, bool]:
        """
        Extract checkbox states from xlsx XML.

        Handles both:
        1. Modern form controls (xl/ctrlProps/)
        2. Legacy VML checkboxes (xl/drawings/vmlDrawing.vml)

        Returns:
            Dict mapping cell addresses to checkbox states (True/False)
        """
        checkboxes = {}
        try:
            with zipfile.ZipFile(xlsx_path, "r") as zf:
                workbook_xml = zf.read("xl/workbook.xml")
                root = ET.fromstring(workbook_xml)
                sheet_idx = None
                for idx, sheet in enumerate(
                    root.findall(f".//{{{NS_SPREADSHEETML}}}sheet"), start=1
                ):
                    if sheet.get("name") == sheet_name:
                        sheet_idx = idx
                        break

                if sheet_idx is None:
                    return checkboxes
                sheet_rels_path = f"xl/worksheets/_rels/sheet{sheet_idx}.xml.rels"
                ctrl_props_map = {}  # rId -> ctrlProps path

                if sheet_rels_path in zf.namelist():
                    rels_xml = zf.read(sheet_rels_path)
                    rels_root = ET.fromstring(rels_xml)
                    for rel in rels_root.findall(".//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
                        rel_type = rel.get("Type", "")
                        if "ctrlProp" in rel_type.lower() or "control" in rel_type.lower():
                            rid = rel.get("Id", "")
                            target = rel.get("Target", "")
                            if target.startswith(".."):
                                ctrl_path = "xl/" + target.replace("../", "")
                            else:
                                ctrl_path = f"xl/worksheets/{target}"
                            ctrl_props_map[rid] = ctrl_path
                sheet_path = f"xl/worksheets/sheet{sheet_idx}.xml"
                if sheet_path in zf.namelist():
                    sheet_xml = zf.read(sheet_path)
                    sheet_root = ET.fromstring(sheet_xml)
                    for control in sheet_root.findall(f".//{{{NS_SPREADSHEETML}}}control"):
                        rid = control.get(f"{{{NS_REL}}}id") or control.get("r:id")
                        if rid and rid in ctrl_props_map:
                            ctrl_path = ctrl_props_map[rid]
                            if ctrl_path in zf.namelist():
                                try:
                                    ctrl_xml = zf.read(ctrl_path)
                                    ctrl_root = ET.fromstring(ctrl_xml)
                                    checked = ctrl_root.get("checked", "").lower() in ("1", "true", "checked")
                                    linked_cell = ctrl_root.get("fmlaLink", "")
                                    if linked_cell:
                                        if "!" in linked_cell:
                                            linked_cell = linked_cell.split("!")[-1]
                                        linked_cell = linked_cell.replace("$", "")
                                        checkboxes[linked_cell] = checked
                                except Exception:
                                    pass

                # legacy VML checkboxes
                vml_path = f"xl/drawings/vmlDrawing{sheet_idx}.vml"
                if vml_path in zf.namelist():
                    vml_content = zf.read(vml_path).decode("utf-8", errors="ignore")
                    checkbox_block_pattern = re.compile(
                        r'<x:ClientData[^>]*ObjectType="Checkbox"[^>]*>(.*?)</x:ClientData>',
                        re.DOTALL | re.IGNORECASE
                    )

                    for block_match in checkbox_block_pattern.finditer(vml_content):
                        block = block_match.group(1)
                        checked_match = re.search(r'<x:Checked>(\d+)</x:Checked>', block)
                        checked = checked_match and checked_match.group(1) == "1"
                        fmla_match = re.search(r'<x:FmlaLink>([^<]+)</x:FmlaLink>', block)
                        if fmla_match:
                            linked_cell = fmla_match.group(1)
                            if "!" in linked_cell:
                                linked_cell = linked_cell.split("!")[-1]
                            linked_cell = linked_cell.replace("$", "")
                            checkboxes[linked_cell] = checked
                        else:
                            anchor_match = re.search(r'<x:Anchor>([^<]+)</x:Anchor>', block)
                            if anchor_match:
                                anchor = anchor_match.group(1)
                                parts = anchor.split(",")
                                if len(parts) >= 4:
                                    try:
                                        col = int(parts[0].strip()) + 1
                                        row = int(parts[2].strip()) + 1
                                        cell_ref = f"{get_column_letter(col)}{row}"
                                        checkboxes[cell_ref] = checked
                                    except (ValueError, IndexError):
                                        pass

        except Exception as exc:
            log.debug("Error extracting checkboxes: %s", exc)

        return checkboxes

    def _extract_images_from_sheet(
        self, ws: Worksheet, xlsx_path: str
    ) -> List[Dict[str, Any]]:
        """
        Extract all images from a worksheet.

        Returns:
            List of image info dicts with keys: path, hash, x, y, width, height,
            anchor_cell, center_row, center_col, formatted_content
        """
        images = []

        if not self.get_image_content:
            return images
        try:
            for img in ws._images:
                try:
                    anchor = img.anchor
                    x, y, width, height = 0, 0, 0, 0
                    anchor_cell = "A1"

                    if isinstance(anchor, TwoCellAnchor):
                        from_col = anchor._from.col
                        from_row = anchor._from.row
                        to_col = anchor.to.col
                        to_row = anchor.to.row
                        anchor_cell = f"{get_column_letter(from_col + 1)}{from_row + 1}"

                        x, y = self._get_cell_pixel_position(ws, from_col + 1, from_row + 1)
                        x += self._emu_to_pixels(anchor._from.colOff or 0)
                        y += self._emu_to_pixels(anchor._from.rowOff or 0)

                        end_x, end_y = self._get_cell_pixel_position(ws, to_col + 1, to_row + 1)
                        end_x += self._emu_to_pixels(anchor.to.colOff or 0)
                        end_y += self._emu_to_pixels(anchor.to.rowOff or 0)

                        width = end_x - x
                        height = end_y - y

                    elif isinstance(anchor, OneCellAnchor):
                        from_col = anchor._from.col
                        from_row = anchor._from.row
                        anchor_cell = f"{get_column_letter(from_col + 1)}{from_row + 1}"

                        x, y = self._get_cell_pixel_position(ws, from_col + 1, from_row + 1)
                        x += self._emu_to_pixels(anchor._from.colOff or 0)
                        y += self._emu_to_pixels(anchor._from.rowOff or 0)

                        width = self._emu_to_pixels(anchor.ext.cx or 0)
                        height = self._emu_to_pixels(anchor.ext.cy or 0)

                    elif isinstance(anchor, AbsoluteAnchor):
                        x = self._emu_to_pixels(anchor.pos.x or 0)
                        y = self._emu_to_pixels(anchor.pos.y or 0)
                        width = self._emu_to_pixels(anchor.ext.cx or 0)
                        height = self._emu_to_pixels(anchor.ext.cy or 0)

                    image_data = None
                    ext = "png"

                    if hasattr(img, "_data") and img._data:
                        image_data = img._data()
                    elif hasattr(img, "ref"):
                        try:
                            with zipfile.ZipFile(xlsx_path, "r") as zf:
                                img_path = img.ref
                                if not img_path.startswith("xl/"):
                                    img_path = f"xl/{img_path}"
                                if img_path in zf.namelist():
                                    image_data = zf.read(img_path)
                                    ext = Path(img_path).suffix.lstrip(".") or "png"
                        except Exception:
                            pass

                    if image_data:
                        saved = self._save_image_bytes(image_data, ext)
                        if saved:
                            path, hash_val = saved
                            bbox = (x, y, x + width, y + height)
                            center_x = x + width / 2
                            center_y = y + height / 2
                            center_row, center_col = self._pixel_to_cell(ws, center_x, center_y)
                            formatted_content = self._format_image_tag(path, hash_val)

                            images.append({
                                "type": "image",
                                "path": path,
                                "hash": hash_val,
                                "x": x,
                                "y": y,
                                "width": width,
                                "height": height,
                                "anchor_cell": anchor_cell,
                                "bbox": bbox,
                                "center_row": center_row,
                                "center_col": center_col,
                                "formatted_content": formatted_content,
                            })

                except Exception as exc:
                    log.debug("Error processing image: %s", exc)

        except Exception as exc:
            log.debug("Error extracting images from sheet: %s", exc)

        return images

    def _extract_charts_from_sheet(
        self, ws: Worksheet, xlsx_path: str
    ) -> List[Dict[str, Any]]:
        """
        Extract charts from worksheet.

        Returns:
            List of chart info dicts with center_row, center_col, formatted_content
        """
        charts = []

        if not self.get_chart_content:
            return charts

        try:
            for chart in ws._charts:
                try:
                    anchor = chart.anchor
                    x, y, width, height = 0, 0, 400, 300  # defuat chart size
                    anchor_cell = "A1"

                    if hasattr(anchor, "_from"):
                        from_col = anchor._from.col
                        from_row = anchor._from.row
                        anchor_cell = f"{get_column_letter(from_col + 1)}{from_row + 1}"
                        x, y = self._get_cell_pixel_position(ws, from_col + 1, from_row + 1)

                    if hasattr(anchor, "to"):
                        end_x, end_y = self._get_cell_pixel_position(
                            ws, anchor.to.col + 1, anchor.to.row + 1
                        )
                        width = end_x - x
                        height = end_y - y

                    bbox = (x, y, x + width, y + height)
                    center_x = x + width / 2
                    center_y = y + height / 2
                    center_row, center_col = self._pixel_to_cell(ws, center_x, center_y)
                    chart_title = getattr(chart, "title", None) or "Chart"
                    formatted_content = f"\n{SHAPE_START_MARKER} Chart: {chart_title} {SHAPE_END_MARKER}\n"

                    charts.append({
                        "type": "chart",
                        "x": x,
                        "y": y,
                        "width": width,
                        "height": height,
                        "anchor_cell": anchor_cell,
                        "bbox": bbox,
                        "title": chart_title,
                        "center_row": center_row,
                        "center_col": center_col,
                        "formatted_content": formatted_content,
                    })

                except Exception as exc:
                    log.debug("Error processing chart: %s", exc)

        except Exception as exc:
            log.debug("Error extracting charts: %s", exc)

        return charts

    def _extract_shapes_from_xml(
        self, xlsx_path: str, sheet_name: str, ws: Worksheet
    ) -> List[Dict[str, Any]]:
        """
        Extract shapes (arrows, diamonds, etc.) from worksheet XML.

        Returns:
            List of shape info dicts
        """
        shapes = []

        if not self.get_shape_content:
            return shapes

        try:
            with zipfile.ZipFile(xlsx_path, "r") as zf:
                workbook_xml = zf.read("xl/workbook.xml")
                root = ET.fromstring(workbook_xml)
                sheet_idx = None
                for idx, sheet in enumerate(
                    root.findall(f".//{{{NS_SPREADSHEETML}}}sheet"), start=1
                ):
                    if sheet.get("name") == sheet_name:
                        sheet_idx = idx
                        break

                if sheet_idx is None:
                    return shapes

                drawing_path = f"xl/drawings/drawing{sheet_idx}.xml"
                if drawing_path not in zf.namelist():
                    sheet_rels_path = f"xl/worksheets/_rels/sheet{sheet_idx}.xml.rels"
                    if sheet_rels_path in zf.namelist():
                        rels_xml = zf.read(sheet_rels_path)
                        rels_root = ET.fromstring(rels_xml)
                        for rel in rels_root.findall(f".//{{{NS_REL}}}Relationship"):
                            if "drawing" in rel.get("Type", "").lower():
                                target = rel.get("Target", "")
                                if target.startswith(".."):
                                    drawing_path = "xl/" + target.replace("../", "")
                                else:
                                    drawing_path = f"xl/worksheets/{target}"
                                break

                if drawing_path not in zf.namelist():
                    return shapes

                drawing_xml = zf.read(drawing_path)
                drawing_root = ET.fromstring(drawing_xml)

                for sp in drawing_root.findall(f".//{{{NS_DRAWING}}}sp"):
                    try:
                        shape_info = self._parse_shape_element(sp, ws)
                        if shape_info:
                            shapes.append(shape_info)
                    except Exception as exc:
                        log.debug("Error parsing shape: %s", exc)

                for cxn in drawing_root.findall(f".//{{{NS_DRAWING}}}cxnSp"):
                    try:
                        shape_info = self._parse_connector_element(cxn, ws)
                        if shape_info:
                            shapes.append(shape_info)
                    except Exception as exc:
                        log.debug("Error parsing connector: %s", exc)

        except Exception as exc:
            log.debug("Error extracting shapes: %s", exc)

        return shapes

    def _parse_shape_element(
        self, sp_element, ws: Worksheet
    ) -> Optional[Dict[str, Any]]:
        """Parse a shape element from drawing XML."""
        try:
            sp_pr = sp_element.find(f".//{{{NS_DRAWING_MAIN}}}spPr")
            if sp_pr is None:
                return None
            xfrm = sp_pr.find(f".//{{{NS_DRAWING_MAIN}}}xfrm")
            if xfrm is None:
                return None

            off = xfrm.find(f"{{{NS_DRAWING_MAIN}}}off")
            ext = xfrm.find(f"{{{NS_DRAWING_MAIN}}}ext")

            if off is None or ext is None:
                return None

            x = self._emu_to_pixels(int(off.get("x", "0")))
            y = self._emu_to_pixels(int(off.get("y", "0")))
            width = self._emu_to_pixels(int(ext.get("cx", "0")))
            height = self._emu_to_pixels(int(ext.get("cy", "0")))

            prst_geom = sp_pr.find(f".//{{{NS_DRAWING_MAIN}}}prstGeom")
            shape_type = "rectangle"
            if prst_geom is not None:
                shape_type = prst_geom.get("prst", "rectangle")

            tx_body = sp_element.find(f".//{{{NS_DRAWING_MAIN}}}txBody")
            text_content = ""
            if tx_body is not None:
                text_parts = []
                for t in tx_body.findall(f".//{{{NS_DRAWING_MAIN}}}t"):
                    if t.text:
                        text_parts.append(t.text)
                text_content = " ".join(text_parts)

            bbox = (x, y, x + width, y + height)

            center_x = x + width / 2
            center_y = y + height / 2
            center_row, center_col = self._pixel_to_cell(ws, center_x, center_y)

            shape_label = text_content if text_content else f"Shape: {shape_type}"                                                                                          
            formatted_content = self._format_shape_tag(shape_label)

            return {
                "type": "shape",
                "shape_type": shape_type,
                "x": x,
                "y": y,
                "width": width,
                "height": height,
                "text": text_content,
                "bbox": bbox,
                "center_row": center_row,
                "center_col": center_col,
                "formatted_content": formatted_content,
            }

        except Exception as exc:
            log.debug("Error parsing shape element: %s", exc)
            return None

    def _parse_connector_element(
        self, cxn_element, ws: Worksheet
    ) -> Optional[Dict[str, Any]]:
        """Parse a connector element from drawing XML."""
        try:
            sp_pr = cxn_element.find(f".//{{{NS_DRAWING_MAIN}}}spPr")
            if sp_pr is None:
                return None

            xfrm = sp_pr.find(f".//{{{NS_DRAWING_MAIN}}}xfrm")
            if xfrm is None:
                return None

            off = xfrm.find(f"{{{NS_DRAWING_MAIN}}}off")
            ext = xfrm.find(f"{{{NS_DRAWING_MAIN}}}ext")

            if off is None or ext is None:
                return None

            x = self._emu_to_pixels(int(off.get("x", "0")))
            y = self._emu_to_pixels(int(off.get("y", "0")))
            width = self._emu_to_pixels(int(ext.get("cx", "0")))
            height = self._emu_to_pixels(int(ext.get("cy", "0")))

            bbox = (x, y, x + width, y + height)

            center_x = x + width / 2
            center_y = y + height / 2
            center_row, center_col = self._pixel_to_cell(ws, center_x, center_y)

            formatted_content = self._format_shape_tag("Connector")

            return {
                "type": "connector",
                "shape_type": "connector",
                "x": x,
                "y": y,
                "width": width,
                "height": height,
                "text": "",
                "bbox": bbox,
                "center_row": center_row,
                "center_col": center_col,
                "formatted_content": formatted_content,
            }

        except Exception as exc:
            log.debug("Error parsing connector: %s", exc)
            return None

    def _cluster_elements(
        self, elements: List[Dict[str, Any]]
    ) -> List[List[Dict[str, Any]]]:
        """
        Cluster nearby elements using DBSCAN.

        Args:
            elements: List of elements with bbox info

        Returns:
            List of clusters, each cluster is a list of elements
        """
        if not elements:
            return []

        if len(elements) == 1:
            return [elements]

        centers = []
        for elem in elements:
            bbox = elem.get("bbox", (0, 0, 0, 0))
            cx = (bbox[0] + bbox[2]) / 2
            cy = (bbox[1] + bbox[3]) / 2
            centers.append([cx, cy])

        centers_array = np.array(centers)

        # Apply DBSCAN
        clustering = DBSCAN(
            eps=self.dbscan_eps, min_samples=self.dbscan_min_samples
        ).fit(centers_array)

        # Group elements by cluster
        clusters: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
        for idx, label in enumerate(clustering.labels_):
            if label == -1:
                # noise point - consider as single cluster
                clusters[len(clusters) + 1000].append(elements[idx])
            else:
                clusters[label].append(elements[idx])

        return list(clusters.values())

    def _get_cluster_bbox(
        self, cluster: List[Dict[str, Any]]
    ) -> Tuple[float, float, float, float]:
        """Get bounding box encompassing all elements in cluster."""
        if not cluster:
            return (0, 0, 0, 0)

        min_x = min(elem["bbox"][0] for elem in cluster)
        min_y = min(elem["bbox"][1] for elem in cluster)
        max_x = max(elem["bbox"][2] for elem in cluster)
        max_y = max(elem["bbox"][3] for elem in cluster)

        return (min_x, min_y, max_x, max_y)

    def _render_excel_sheet_to_image(
        self,
        xlsx_path: str,
        sheet_name: str
    ) -> Optional[Image.Image]:
        """
        Render Excel sheet to image using LibreOffice → PDF → Image conversion.

        Args:
            xlsx_path: Path to Excel file
            sheet_name: Name of the sheet to render

        Returns:
            PIL Image object or None if failed
        """
        temp_pdf = None
        try:
            output_dir = tempfile.mkdtemp(prefix="xlsx_pdf_")

            try:
                result = subprocess.run(
                    [
                        "libreoffice",
                        "--headless",
                        "--convert-to",
                        "pdf",
                        "--outdir",
                        output_dir,
                        xlsx_path,
                    ],
                    check=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    timeout=30,
                )
            except FileNotFoundError:
                log.warning("LibreOffice not found. Cannot render Excel sheet.")
                return None
            except subprocess.TimeoutExpired:
                log.warning("LibreOffice conversion timeout.")
                return None
            except subprocess.CalledProcessError as exc:
                log.warning("LibreOffice conversion failed: %s", exc.stderr)
                return None

            pdf_filename = os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf"
            temp_pdf = os.path.join(output_dir, pdf_filename)

            if not os.path.exists(temp_pdf):
                log.warning("PDF file not generated at %s", temp_pdf)
                return None

            try:

                # Convert PDF to image using pdf2image (primary method)
                images = convert_from_path(temp_pdf, dpi=200, first_page=1, last_page=1)

                if images:
                    return images[0]
                else:
                    log.warning("No images generated from PDF")
                    return None

            except ImportError:
                log.error("pdf2image not available. Please install: pip install pdf2image")
                log.error("Also ensure poppler-utils is installed: apt-get install poppler-utils")
                return None
            except Exception as exc:
                log.warning("pdf2image conversion failed: %s", exc)
                return None

        except Exception as exc:
            log.warning("Error rendering Excel sheet: %s", exc)
            return None
        finally:
            if temp_pdf and os.path.exists(temp_pdf):
                try:
                    os.remove(temp_pdf)
                except:
                    pass
            try:
                if 'output_dir' in locals() and os.path.exists(output_dir):
                    import shutil
                    shutil.rmtree(output_dir)
            except:
                pass

    def _create_cluster_image(
        self,
        ws: Worksheet,
        xlsx_path: str,
        cluster: List[Dict[str, Any]],
        bbox: Tuple[float, float, float, float]
    ) -> Tuple[str, str]:
        """
        Crop Excel sheet content at cluster bbox by rendering sheet to image then cropping.

        Process:
        1. Render Excel sheet to image (LibreOffice → PDF → Image)
        2. Crop image at bbox coordinates
        3. Save cropped region

        Args:
            ws: Worksheet object
            xlsx_path: Path to Excel file
            cluster: List of elements in the cluster
            bbox: Bounding box of the cluster (min_x, min_y, max_x, max_y)

        Returns:
            Tuple of (image_path, image_hash)
        """
        try:
            sheet_name = ws.title

            sheet_image = self._render_excel_sheet_to_image(xlsx_path, sheet_name)

            if sheet_image is None:
                cluster_images = [e for e in cluster if e.get("type") == "image" and e.get("path")]
                if cluster_images:
                    largest_image = max(cluster_images, key=lambda x: x.get("width", 0) * x.get("height", 0))
                    return (largest_image["path"], largest_image["hash"])
                else:
                    width = max(int(bbox[2] - bbox[0]), 100)
                    height = max(int(bbox[3] - bbox[1]), 100)
                    placeholder = Image.new('RGB', (width, height), color='white')
                    img_buffer = io.BytesIO()
                    placeholder.save(img_buffer, format='PNG')
                    img_bytes = img_buffer.getvalue()
                    result = self._save_image_bytes(img_bytes, ext='png', name_hint='cluster_placeholder')
                    return result if result else ("", "")
            crop_box = (
                max(0, int(bbox[0])),
                max(0, int(bbox[1])),
                min(sheet_image.width, int(bbox[2])),
                min(sheet_image.height, int(bbox[3]))
            )

            if crop_box[0] >= crop_box[2] or crop_box[1] >= crop_box[3]:
                log.warning("Invalid crop box: %s", crop_box)
                cropped_image = sheet_image
            else:
                cropped_image = sheet_image.crop(crop_box)

            img_buffer = io.BytesIO()
            cropped_image.save(img_buffer, format='PNG')
            img_bytes = img_buffer.getvalue()

            result = self._save_image_bytes(
                img_bytes,
                ext='png',
                name_hint=f'cluster_crop_{len(cluster)}_elements'
            )

            if result:
                return result
            else:
                return ("", "")

        except Exception as exc:
            log.warning("Error creating cluster image: %s", exc)
            return ("", "")

    def _extract_cell_data(
        self, ws: Worksheet, checkboxes: Dict[str, bool]
    ) -> Dict[str, Any]:
        """
        Extract all cell data from worksheet.

        Returns:
            Dict with 'data' (2D array), 'merged_cells', 'borders'
        """
        min_row = ws.min_row or 1
        max_row = ws.max_row or 1
        min_col = ws.min_column or 1
        max_col = ws.max_column or 1
        data = []
        borders = {}  

        for row_idx in range(min_row, max_row + 1):
            row_data = []
            for col_idx in range(min_col, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_ref = f"{get_column_letter(col_idx)}{row_idx}"

                if cell_ref in checkboxes:
                    row_data.append(checkboxes[cell_ref])
                else:
                    value = cell.value
                    if value is None:
                        row_data.append("")
                    elif isinstance(value, bool):
                        row_data.append(value)
                    else:
                        row_data.append(str(value))

                if cell.border:
                    border = cell.border
                    borders[cell_ref] = {
                        "left": border.left.style if border.left else None,
                        "right": border.right.style if border.right else None,
                        "top": border.top.style if border.top else None,
                        "bottom": border.bottom.style if border.bottom else None,
                    }

            data.append(row_data)

        merged_cells = []
        for merged_range in ws.merged_cells.ranges:
            merged_cells.append({
                "range": str(merged_range),
                "min_col": merged_range.min_col,
                "max_col": merged_range.max_col,
                "min_row": merged_range.min_row,
                "max_row": merged_range.max_row,
            })

        return {
            "data": data,
            "merged_cells": merged_cells,
            "borders": borders,
            "min_row": min_row,
            "max_row": max_row,
            "min_col": min_col,
            "max_col": max_col,
        }

    def _has_full_border(self, borders: Dict[str, Any], cell_ref: str) -> bool:
        """Check if a cell has full border (all four sides)."""
        if cell_ref not in borders:
            return False
        border = borders[cell_ref]
        return all([
            border.get("left"),
            border.get("right"),
            border.get("top"),
            border.get("bottom"),
        ])

    def _has_border(
        self,
        borders: Dict[str, Any],
        row: int,
        col: int,
        side: str,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int
    ) -> bool:
        """
        Check if a cell has a border on a given side.

        Args:
            borders: Dictionary of cell borders
            row: Row index
            col: Column index
            side: Border side ('top', 'bottom', 'left', 'right')
            min_row: Minimum row boundary
            max_row: Maximum row boundary
            min_col: Minimum column boundary
            max_col: Maximum column boundary

        Returns:
            True if cell has border on the specified side, False otherwise
        """
        if row < min_row or row > max_row or col < min_col or col > max_col:
            return False
        cell_ref = f"{get_column_letter(col)}{row}"
        if cell_ref not in borders:
            return False
        return borders[cell_ref].get(side) is not None

    def _is_top_left_corner(
        self,
        borders: Dict[str, Any],
        row: int,
        col: int,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int
    ) -> bool:
        """
        Check if a cell is a top-left corner (has both top and left borders).

        Args:
            borders: Dictionary of cell borders
            row: Row index
            col: Column index
            min_row: Minimum row boundary
            max_row: Maximum row boundary
            min_col: Minimum column boundary
            max_col: Maximum column boundary

        Returns:
            True if cell has both top and left borders, False otherwise
        """
        return (
            self._has_border(borders, row, col, 'top', min_row, max_row, min_col, max_col) and
            self._has_border(borders, row, col, 'left', min_row, max_row, min_col, max_col)
        )

    def _get_next_border_direction(
        self,
        current_direction: str,
        current_row: int,
        current_col: int,
        has_border_fn
    ) -> str:
        """
        Determine the next direction when tracing table borders.

        Args:
            current_direction: Current movement direction ('right', 'down', 'left', 'up')
            current_row: Current row position
            current_col: Current column position
            has_border_fn: Function that checks if a cell has a border on a given side

        Returns:
            New direction to move in
        """
        if current_direction == 'right':
            if has_border_fn(current_row - 1, current_col + 1, 'left') or has_border_fn(current_row - 1, current_col, 'right'):
                return 'up!'
            if has_border_fn(current_row, current_col + 1, 'top') or has_border_fn(current_row - 1, current_col + 1, 'bottom'):
                return 'right'  # Keep moving right
            if has_border_fn(current_row, current_col, 'right') or has_border_fn(current_row, current_col + 1, 'left'):
                return 'down'
        elif current_direction == 'down':
            if has_border_fn(current_row + 1, current_col + 1, 'top') or has_border_fn(current_row - 1, current_col + 1, 'bottom'):
                return 'right!'
            if has_border_fn(current_row + 1, current_col, 'right') or has_border_fn(current_row + 1, current_col + 1, 'left'):
                return 'down'  # Keep moving down
            if has_border_fn(current_row, current_col, 'bottom') or has_border_fn(current_row + 1, current_col, 'top'):
                return 'left'
        elif current_direction == 'left':
            if has_border_fn(current_row + 1, current_col - 1, 'right') or has_border_fn(current_row + 1, current_col, 'left'):
                return 'down!'
            if has_border_fn(current_row, current_col - 1, 'bottom') or has_border_fn(current_row + 1, current_col - 1, 'top'):
                return 'left'  # Keep moving left
            if has_border_fn(current_row, current_col, 'left') or has_border_fn(current_row, current_col - 1, 'right'):
                return 'up'
        elif current_direction == 'up':
            if has_border_fn(current_row - 1, current_col - 1, 'bottom') or has_border_fn(current_row, current_col - 1, 'top'):
                return 'left!'
            if has_border_fn(current_row - 1, current_col, 'left') or has_border_fn(current_row - 1, current_col - 1, 'right'):
                return 'up'  # Keep moving up
            if has_border_fn(current_row, current_col, 'top') or has_border_fn(current_row - 1, current_col, 'bottom'):
                return 'right'

        return current_direction  # Default: keep same direction

    def _find_table_regions(
        self, cell_data: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """
        Find table regions using boundary-tracing algorithm.

        Algorithm:
        1. Find first cell with (top border OR colored background) AND (left border OR colored background) - top-left corner
        2. Start moving right direction
        3. At each cell, decide next direction based on adjacent borders OR colored background:
           - Moving RIGHT: if top cell has (left border OR colored background) → turn UP
                          else if bottom cell has (right border OR colored background) → turn DOWN
                          else → continue RIGHT
           - Similar logic for other directions (DOWN, LEFT, UP)
        4. Continue until returning to start cell
        5. Extract table region from corner cells (where direction changed)

        Returns:
            List of table regions with bounds and data
        """
        tables = []
        borders = cell_data["borders"]
        data = cell_data["data"]
        min_row = cell_data["min_row"]
        min_col = cell_data["min_col"]
        max_row = cell_data["max_row"]
        max_col = cell_data["max_col"]
        visited_starts = set()

        has_border = lambda row, col, side: self._has_border(
            borders, row, col, side, min_row, max_row, min_col, max_col
        )
        is_top_left_corner = lambda row, col: self._is_top_left_corner(
            borders, row, col, min_row, max_row, min_col, max_col
        )

        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                if (row_idx, col_idx) in visited_starts:
                    continue

                if not is_top_left_corner(row_idx, col_idx):
                    continue

            
                start_row, start_col = row_idx, col_idx
                current_row, current_col = start_row, start_col
                direction = 'right'
                corners = [(current_row, current_col)]  
                visited_cells = {(current_row, current_col)}
                max_iterations = (max_row - min_row + 2) * (max_col - min_col + 2) * 4
                iterations = 0

                while iterations < max_iterations:
                    iterations += 1
                    prev_direction = direction

                    direction = self._get_next_border_direction(
                        direction, current_row, current_col, has_border
                    )
                
                    is_counter = False
                    if direction[-1] == '!':
                        is_counter = True
                        direction = direction[:-1]
                    if direction != prev_direction:
                        if is_counter:
                            if prev_direction == 'right':
                                current_row, current_col = current_row, current_col + 1
                            elif prev_direction == 'down':
                                current_row, current_col = current_row + 1, current_col
                            elif prev_direction == 'left':
                                current_row, current_col = current_row, current_col - 1
                            elif prev_direction == 'up':
                                current_row, current_col = current_row - 1, current_col
                        
                        corners.append((current_row, current_col))

                    if direction == prev_direction or is_counter: 
                        if direction == 'right':
                            next_row, next_col = current_row, current_col + 1
                        elif direction == 'down':
                            next_row, next_col = current_row + 1, current_col
                        elif direction == 'left':
                            next_row, next_col = current_row, current_col - 1
                        elif direction == 'up':
                            next_row, next_col = current_row - 1, current_col
                    else:
                        next_row, next_col = current_row, current_col
                    if (next_row, next_col) == (start_row, start_col):
                        break

                    if next_row < min_row or next_row > max_row or next_col < min_col or next_col > max_col:
                        break

                    visited_cells.add((next_row, next_col))
                    current_row, current_col = next_row, next_col

                if len(corners) >= 3:
                    corner_rows = [c[0] for c in corners]
                    corner_cols = [c[1] for c in corners]
                    table_min_row = min(corner_rows)
                    table_max_row = max(corner_rows)
                    table_min_col = min(corner_cols)
                    table_max_col = max(corner_cols)

                    for r in range(table_min_row, table_max_row + 1):
                        for c in range(table_min_col, table_max_col + 1):
                            visited_starts.add((r, c))

                    table_data = []
                    for r in range(table_min_row, table_max_row + 1):
                        row_data = []
                        for c in range(table_min_col, table_max_col + 1):
                            data_row = r - min_row
                            data_col = c - min_col
                            if 0 <= data_row < len(data) and 0 <= data_col < len(data[data_row]):
                                val = data[data_row][data_col]
                                row_data.append(str(val) if val else "")
                            else:
                                row_data.append("")
                        table_data.append(row_data)

                    if len(table_data) > 0 and len(table_data[0]) > 0:
                        tables.append({
                            "min_row": table_min_row,
                            "max_row": table_max_row,
                            "min_col": table_min_col,
                            "max_col": table_max_col,
                            "data": table_data,
                            "corners": corners, 
                        })
        return tables

    def _find_nested_tables_in_region(
        self,
        cell_data: Dict[str, Any],
        parent_table: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """
        Find nested tables within a parent table region.

        Algorithm:
        1. Loop through cells in parent table region
        2. Find cells with (top border AND left border) that are NOT in first row/column
        3. Use same border-tracking algorithm as first-level tables to trace boundaries

        Args:
            cell_data: Cell data dict with borders, data, etc.
            parent_table: Parent table region dict

        Returns:
            List of nested table regions with bounds and data
        """
        nested_tables = []
        borders = cell_data["borders"]
        data = cell_data["data"]
        min_row = cell_data["min_row"]
        min_col = cell_data["min_col"]
        max_row = cell_data["max_row"]
        max_col = cell_data["max_col"]

        parent_min_row = parent_table["min_row"]
        parent_max_row = parent_table["max_row"]
        parent_min_col = parent_table["min_col"]
        parent_max_col = parent_table["max_col"]

        visited_starts = set()

        has_border = lambda row, col, side: self._has_border(
            borders, row, col, side, min_row, max_row, min_col, max_col
        )
        for row_idx in range(parent_min_row, parent_max_row + 1):
            for col_idx in range(parent_min_col, parent_max_col + 1):
                if (row_idx, col_idx) in visited_starts:
                    continue
                if row_idx == parent_min_row or col_idx == parent_min_col: 
                    continue

                if has_border(row_idx - 1, col_idx, 'left') or has_border(row_idx, col_idx - 1, 'top'):
                    continue

                if not (has_border(row_idx, col_idx, 'top') and has_border(row_idx, col_idx, 'left')):
                    continue
                
                start_row, start_col = row_idx, col_idx

                current_row, current_col = start_row, start_col
                direction = 'right'
                corners = [(current_row, current_col)]
                visited_cells = {(current_row, current_col)}

                max_iterations = (parent_max_row - parent_min_row + 2) * (parent_max_col - parent_min_col + 2) * 4
                iterations = 0

                while iterations < max_iterations:
                    iterations += 1
                    prev_direction = direction

                    direction = self._get_next_border_direction(
                        direction, current_row, current_col, has_border
                    )
                    is_counter = False
                    if direction[-1] == '!':
                        is_counter = True
                        direction = direction[:-1]
                    if direction != prev_direction:
                        if is_counter:
                            if prev_direction == 'right':
                                current_row, current_col = current_row, current_col + 1
                            elif prev_direction == 'down':
                                current_row, current_col = current_row + 1, current_col
                            elif prev_direction == 'left':
                                current_row, current_col = current_row, current_col - 1
                            elif prev_direction == 'up':
                                current_row, current_col = current_row - 1, current_col
                        
                        corners.append((current_row, current_col))

                    if direction == prev_direction or is_counter: 
                        if direction == 'right':
                            next_row, next_col = current_row, current_col + 1
                        elif direction == 'down':
                            next_row, next_col = current_row + 1, current_col
                        elif direction == 'left':
                            next_row, next_col = current_row, current_col - 1
                        elif direction == 'up':
                            next_row, next_col = current_row - 1, current_col
                    else:
                        next_row, next_col = current_row, current_col
                    
                    if (next_row, next_col) == (start_row, start_col):
                        break

                    if (next_row < parent_min_row or next_row > parent_max_row or
                        next_col < parent_min_col or next_col > parent_max_col):
                        break


                    visited_cells.add((next_row, next_col))
                    current_row, current_col = next_row, next_col

                if len(corners) >= 3:
                    corner_rows = [c[0] for c in corners]
                    corner_cols = [c[1] for c in corners]
                    table_min_row = min(corner_rows)
                    table_max_row = max(corner_rows)
                    table_min_col = min(corner_cols)
                    table_max_col = max(corner_cols)

                    for r in range(table_min_row, table_max_row + 1):
                        for c in range(table_min_col, table_max_col + 1):
                            visited_starts.add((r, c))

                    table_data = []
                    for r in range(table_min_row, table_max_row + 1):
                        row_data = []
                        for c in range(table_min_col, table_max_col + 1):
                            data_row = r - min_row
                            data_col = c - min_col
                            if 0 <= data_row < len(data) and 0 <= data_col < len(data[data_row]):
                                val = data[data_row][data_col]
                                row_data.append(str(val) if val else "")
                            else:
                                row_data.append("")
                        table_data.append(row_data)

                    if len(table_data) > 0 and len(table_data[0]) > 0:
                        nested_tables.append({
                            "min_row": table_min_row,
                            "max_row": table_max_row,
                            "min_col": table_min_col,
                            "max_col": table_max_col,
                            "data": table_data,
                            "corners": corners,
                        })

        return nested_tables

    def _merge_cell_contents_into_table(
        self,
        table: Dict[str, Any],
        cell_contents: Dict[Tuple[int, int], str]
    ) -> None:
        """
        Merge visual content (images, charts, shapes) from cell_contents into table data.

        Args:
            table: Table dict with min_row, max_row, min_col, max_col, data
            cell_contents: Dict mapping (row, col) -> visual content string
        """
        table_data = table["data"]
        min_row = table["min_row"]
        min_col = table["min_col"]

        for (row, col), visual_content in cell_contents.items():
            if table["min_row"] <= row <= table["max_row"] and table["min_col"] <= col <= table["max_col"]:
                data_row_idx = row - min_row
                data_col_idx = col - min_col

                if 0 <= data_row_idx < len(table_data) and 0 <= data_col_idx < len(table_data[data_row_idx]):
                    existing_content = table_data[data_row_idx][data_col_idx]
                    if existing_content and existing_content.strip():
                        table_data[data_row_idx][data_col_idx] = existing_content + "\n" + visual_content
                    else:
                        table_data[data_row_idx][data_col_idx] = visual_content
                    

    def _format_table_with_nested_tables(
        self,
        cell_data: Dict[str, Any],
        parent_table: Dict[str, Any]
    ) -> List[List[str]]:
        """
        Format a parent table, detecting nested tables and replacing their cells
        with markdown-tagged content.

        Args:
            cell_data: Cell data dict
            parent_table: Parent table dict

        Returns:
            Modified table data with nested tables replaced by markdown tags
        """
        nested_tables = self._find_nested_tables_in_region(cell_data, parent_table)

        table_data = [row[:] for row in parent_table["data"]] # deep copy
        for nested_table in nested_tables:
            nested_markdown = self._table_to_markdown(nested_table["data"])

            if not nested_markdown:
                continue

            nested_content = f"{TABLE_START_MARKER} {nested_markdown} {TABLE_END_MARKER}"
            nested_min_row = nested_table["min_row"]
            nested_min_col = nested_table["min_col"]
            nested_max_row = nested_table["max_row"]
            nested_max_col = nested_table["max_col"]

            parent_min_row = parent_table["min_row"]
            parent_min_col = parent_table["min_col"]

            top_left_row_idx = nested_min_row - parent_min_row
            top_left_col_idx = nested_min_col - parent_min_col
            
            if 0 <= top_left_row_idx < len(table_data) and 0 <= top_left_col_idx < len(table_data[top_left_row_idx]):
                table_data[top_left_row_idx][top_left_col_idx] = nested_content

            for row_offset in range(nested_max_row - nested_min_row + 1):
                for col_offset in range(nested_max_col - nested_min_col + 1):
                    if row_offset == 0 and col_offset == 0:
                        continue

                    data_row_idx = top_left_row_idx + row_offset
                    data_col_idx = top_left_col_idx + col_offset

                    if (0 <= data_row_idx < len(table_data) and
                        0 <= data_col_idx < len(table_data[data_row_idx])):
                        table_data[data_row_idx][data_col_idx] = ""

        return table_data

    def _find_heading_cell(
        self, cell_data: Dict[str, Any], tables: List[Dict[str, Any]]
    ) -> Optional[str]:
        """
        Find potential heading cell for the sheet content.

        Look for:
        1. Highest (lowest row index) non-empty cell without full border
        2. Cell containing "表" followed by numeric text
        """
        borders = cell_data["borders"]
        data = cell_data["data"]
        min_row = cell_data["min_row"]
        min_col = cell_data["min_col"]
        max_row = cell_data["max_row"]
        max_col = cell_data["max_col"]

        table_cells = set()
        for table in tables:
            for r in range(table["min_row"], table["max_row"] + 1):
                for c in range(table["min_col"], table["max_col"] + 1):
                    table_cells.add((r, c))

        heading_text = ""

        # Strategy 1: Find highest non-empty, non-bordered cell
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                if (row_idx, col_idx) in table_cells:
                    continue

                cell_ref = f"{get_column_letter(col_idx)}{row_idx}"

                if self._has_full_border(borders, cell_ref):
                    continue
                data_row = row_idx - min_row
                data_col = col_idx - min_col
                if 0 <= data_row < len(data) and 0 <= data_col < len(data[data_row]):
                    val = data[data_row][data_col]
                    if val and str(val).strip():
                        heading_text = str(val).strip()
                        return heading_text

        # Strategy 2: Look for "表" pattern
        table_pattern = re.compile(r"表\s*\d+")
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                if (row_idx, col_idx) in table_cells:
                    continue

                cell_ref = f"{get_column_letter(col_idx)}{row_idx}"

                if self._has_full_border(borders, cell_ref):
                    continue

                data_row = row_idx - min_row
                data_col = col_idx - min_col
                if 0 <= data_row < len(data) and 0 <= data_col < len(data[data_row]):
                    val = data[data_row][data_col]
                    if val and table_pattern.search(str(val)):
                        return str(val).strip()

        return heading_text if heading_text else ""

    def _table_to_markdown(self, table_data: List[List[str]]) -> str:
        """Convert table data to markdown format."""
        if not table_data:
            return ""

        max_cols = max(len(row) for row in table_data) if table_data else 0
        normalized = [row + [""] * (max_cols - len(row)) for row in table_data]

        if not normalized:
            return ""
        header = normalized[0]
        separator = ["---"] * len(header)

        lines = [
            "| " + " | ".join(cell.replace("|", "\\|") for cell in header) + " |",
            "| " + " | ".join(separator) + " |",
        ]

        for row in normalized[1:]:
            lines.append(
                "| " + " | ".join(cell.replace("|", "\\|") for cell in row) + " |"
            )

        return "\n".join(lines)

    def _format_image_tag(self, path: str, hash_val: str) -> str:
        """Format image path tag consistent with docx_reader."""
        return f"\n{IMAGE_PATH_START_MARKER} {path}|{hash_val} {IMAGE_PATH_END_MARKER}\n"

    def _format_shape_tag(self, text: str) -> str:
        """Format shape content tag."""
        return f"\n{SHAPE_START_MARKER} {text} {SHAPE_END_MARKER}\n"

    def _format_table_tag(self, markdown: str) -> str:
        """Format table content tag."""
        return f"\n{TABLE_START_MARKER} {markdown} {TABLE_END_MARKER}\n"

    def _extract_sheet_data(
        self, ws: Worksheet, xlsx_path: str, sheet_name: str
    ) -> Dict[str, Any]:
        """
        Step 1: Extract all data from a single sheet.

        Returns:
            Dict containing table_data, images, charts, shapes, merged_cells,
            cell_contents (with visual elements inserted), outside_table_cells
        """
        
        checkboxes = self._extract_checkboxes_from_xml(xlsx_path, sheet_name)
        cell_data = self._extract_cell_data(ws, checkboxes)
        images = self._extract_images_from_sheet(ws, xlsx_path)
        charts = self._extract_charts_from_sheet(ws, xlsx_path)
        shapes = self._extract_shapes_from_xml(xlsx_path, sheet_name, ws)
        all_visual_elements = images + charts + shapes
        clusters = self._cluster_elements(all_visual_elements)
        processed_elements = []
        for cluster in clusters:
            if len(cluster) == 1:
                processed_elements.append(cluster[0])
            else:
                bbox = self._get_cluster_bbox(cluster)
                cluster_image_path, cluster_hash = self._create_cluster_image(
                    ws, xlsx_path, cluster, bbox
                )
                center_x = (bbox[0] + bbox[2]) / 2
                center_y = (bbox[1] + bbox[3]) / 2
                center_row, center_col = self._pixel_to_cell(ws, center_x, center_y)
                formatted_content = self._format_image_tag(cluster_image_path, cluster_hash)

                processed_elements.append({
                    "type": "image", 
                    "bbox": bbox,
                    "x": bbox[0],
                    "y": bbox[1],
                    "width": bbox[2] - bbox[0],
                    "height": bbox[3] - bbox[1],
                    "path": cluster_image_path,
                    "hash": cluster_hash,
                    "center_row": center_row,
                    "center_col": center_col,
                    "formatted_content": formatted_content,
                })

        tables = self._find_table_regions(cell_data)
        if len(tables) > 1:
            layout_type = "table"
        elif len(tables) == 1:
            single_table = tables[0]
            layout_type = self._detect_sheet_layout_type_for_table(cell_data, single_table)
            if layout_type == "zone":
                tables = self._divide_zone_layout_into_tables(cell_data, single_table)
        else:
            layout_type = "table"

        cell_contents: Dict[Tuple[int, int], str] = {}

        for elem in processed_elements:
            center_row = elem.get("center_row")
            center_col = elem.get("center_col")
            formatted_content = elem.get("formatted_content", "")

            if center_row and center_col and formatted_content:
                key = (center_row, center_col)
                if key in cell_contents:
                    cell_contents[key] += formatted_content
                else:
                    cell_contents[key] = formatted_content

        for table in tables:
            self._merge_cell_contents_into_table(table, cell_contents)
            
        table_cells = set()
        for table in tables:
            for r in range(table["min_row"], table["max_row"] + 1):
                for c in range(table["min_col"], table["max_col"] + 1):
                    table_cells.add((r, c))

        outside_table_cells = []
        data = cell_data["data"]
        min_row = cell_data["min_row"]
        min_col = cell_data["min_col"]
        max_row = cell_data["max_row"]
        max_col = cell_data["max_col"]

        heading_text = self._find_heading_cell(cell_data, tables)

        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                if (row_idx, col_idx) in table_cells:
                    continue

                data_row = row_idx - min_row
                data_col = col_idx - min_col

                if 0 <= data_row < len(data) and 0 <= data_col < len(data[data_row]):
                    val = data[data_row][data_col]
                    if val and str(val).strip():
                        cell_text = str(val).strip()
                        if cell_text == heading_text:
                            continue
                        visual_content = cell_contents.get((row_idx, col_idx), "")

                        outside_table_cells.append({
                            "row": row_idx,
                            "col": col_idx,
                            "text": cell_text,
                            "visual_content": visual_content,
                            "cell_ref": f"{get_column_letter(col_idx)}{row_idx}",
                        })

        for (row, col), content in cell_contents.items():
            if (row, col) not in table_cells:
                exists = any(c["row"] == row and c["col"] == col for c in outside_table_cells)
                if not exists:
                    outside_table_cells.append({
                        "row": row,
                        "col": col,
                        "text": "",
                        "visual_content": content,
                        "cell_ref": f"{get_column_letter(col)}{row}",
                    })

        return {
            "sheet_name": sheet_name,
            "cell_data": cell_data,
            "tables": tables,
            "images": images,
            "charts": charts,
            "shapes": shapes,
            "visual_elements": processed_elements,
            "checkboxes": checkboxes,
            "cell_contents": cell_contents,
            "outside_table_cells": outside_table_cells,
            "heading_text": heading_text,
            "layout_type": layout_type,
        }

    def _get_table_center(self, table: Dict[str, Any]) -> Tuple[float, float]:
        """
        Calculate the center point of a table.

        Args:
            table: Table dict with min_row, max_row, min_col, max_col

        Returns:
            Tuple of (center_row, center_col) as floats
        """
        center_row = (table["min_row"] + table["max_row"]) / 2.0
        center_col = (table["min_col"] + table["max_col"]) / 2.0
        return center_row, center_col

    def _calculate_distance(
        self, row1: float, col1: float, row2: float, col2: float
    ) -> float:
        """
        Calculate Euclidean distance between two points.

        Args:
            row1, col1: First point
            row2, col2: Second point

        Returns:
            Distance as float
        """
        return ((row1 - row2) ** 2 + (col1 - col2) ** 2) ** 0.5

    def _position_relative_to_table(
        self, cell_row: int, cell_col: int, table: Dict[str, Any]
    ) -> str:
        """
        Determine if a cell is before or after a table.

        Rules:
        - if row < table_min_row → "before"
        - if row in table row range:
            - if col < table_min_col → "before"
            - if col > table_max_col → "after"
        - if row > table_max_row → "after"

        Returns:
            "before", "after", or "inside"
        """
        table_min_row = table["min_row"]
        table_max_row = table["max_row"]
        table_min_col = table["min_col"]
        table_max_col = table["max_col"]

        if cell_row < table_min_row:
            return "before"
        elif cell_row > table_max_row:
            return "after"
        else:
            if cell_col < table_min_col:
                return "before"
            elif cell_col > table_max_col:
                return "after"
            else:
                return "inside"

    def _build_sheet_tree(self, sheet_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Step 2: Build document tree for a single sheet.

        Assigns each content item to its nearest table using intelligent distance-based algorithm.
        Since tables and content are sorted by position, we optimize by checking if distance
        to next table is shorter and breaking when distance starts increasing.

        Returns:
            Tree structure with heading_text, heading_level, content, children
        """
        sheet_name = sheet_data["sheet_name"]
        layout_type = sheet_data["layout_type"]
        tables = sheet_data["tables"]
        cell_data = sheet_data["cell_data"]
        outside_table_cells = sheet_data.get("outside_table_cells", [])
        heading_text = sheet_data.get("heading_text", "")

        sheet_node = {
            "heading_text": sheet_name,
            "heading_level": 1,
            "content": "",
            "children": [],
        }

        children = []

        if not tables:
            if outside_table_cells:
                sorted_cells = sorted(outside_table_cells, key=lambda x: (x["row"], x["col"]))

                content_parts = []
                for cell_info in sorted_cells:
                    cell_text = cell_info.get("text", "")
                    visual_content = cell_info.get("visual_content", "")

                    combined = ""
                    if cell_text:
                        combined += cell_text
                    if visual_content:
                        combined += visual_content

                    if combined:
                        content_parts.append(combined)

                if content_parts:
                    child_node = {
                        "heading_text": heading_text or "",
                        "heading_level": 2,
                        "content": "\n".join(content_parts),
                        "children": [],
                    }
                    children.append(child_node)
        else:
            table_centers = []
            for table in tables:
                center_row, center_col = self._get_table_center(table)
                table_centers.append((center_row, center_col))
            sorted_content = sorted(outside_table_cells, key=lambda x: (x["row"], x["col"]))

            table_content_map = {i: [] for i in range(len(tables))}
            nearest_table_idx = 0

            for cell_info in sorted_content:
                cell_row = cell_info["row"]
                cell_col = cell_info["col"]
                cell_text = cell_info.get("text", "")
                visual_content = cell_info.get("visual_content", "")
                combined_content = ""
                if cell_text:
                    combined_content += cell_text
                if visual_content:
                    combined_content += visual_content

                if not combined_content:
                    continue

                min_distance = self._calculate_distance(
                    cell_row, cell_col,
                    table_centers[nearest_table_idx][0], table_centers[nearest_table_idx][1]
                )
                for table_idx in range(nearest_table_idx, len(tables)):
                    dist = self._calculate_distance(
                        cell_row, cell_col,
                        table_centers[table_idx][0], table_centers[table_idx][1]
                    )

                    if dist <= min_distance:
                        min_distance = dist
                        nearest_table_idx = table_idx
                    else:
                        break
                table_content_map[nearest_table_idx].append({
                    "row": cell_row,
                    "col": cell_col,
                    "content": combined_content,
                })

            for table_idx, table in enumerate(tables):
                assigned_content = table_content_map[table_idx]

                before_content_parts = []
                after_content_parts = []

                for content_item in assigned_content:
                    position = self._position_relative_to_table(
                        content_item["row"], content_item["col"], table
                    )

                    if position == "before":
                        before_content_parts.append(content_item)
                    elif position == "after":
                        after_content_parts.append(content_item)

                before_content_parts.sort(key=lambda x: (x["row"], x["col"]))
                after_content_parts.sort(key=lambda x: (x["row"], x["col"]))
                content_parts = []

                for part in before_content_parts:
                    content_parts.append(part["content"])

                # Check if this is a zone or table layout
                table_layout_type = table.get("layout_type", layout_type)

                if table_layout_type == "zone":
                    # For zone layout: just output the text directly
                    zone_text = table.get("text", "")
                    if zone_text:
                        content_parts.append(zone_text)
                else:
                    # For table layout: process with header detection and markdown formatting
                    table_data = table.get("data", [])
                    if table_data:
                        # Remove empty rows
                        self._remove_empty_rows(table)
                        table_data = table.get("data", [])

                        if table_data:
                            # Detect header rows
                            header_row_indices = self._detect_table_header_rows(table_data)
                            print(f'[DEBUG] Header rows: {header_row_indices}')

                            # Separate header rows and data rows
                            header_rows = [table_data[i] for i in header_row_indices if i < len(table_data)]
                            data_row_indices = [i for i in range(len(table_data)) if i not in header_row_indices]

                            if data_row_indices:
                                data_rows = [table_data[i] for i in data_row_indices]

                                # Fill forward header rows
                                filled_header_rows = self._forward_fill_table(header_rows)

                                # Fill forward data rows
                                filled_data_rows = self._forward_fill_table(data_rows)

                                # Format as markdown: header rows + data rows
                                markdown_parts = []

                                # Add header rows
                                for header_row in filled_header_rows:
                                    header_str = " | ".join(str(cell) for cell in header_row)
                                    markdown_parts.append(f"| {header_str} |")

                                # Add separator
                                if filled_header_rows:
                                    separator = " | ".join(["---"] * len(filled_header_rows[-1]))
                                    markdown_parts.append(f"| {separator} |")

                                # Add data rows
                                for data_row in filled_data_rows:
                                    data_str = " | ".join(str(cell) for cell in data_row)
                                    markdown_parts.append(f"| {data_str} |")

                                table_markdown = "\n".join(markdown_parts)
                                formatted_table_data = self._format_table_tag(table_markdown)
                                content_parts.append(formatted_table_data)

                for part in after_content_parts:
                    content_parts.append(part["content"])

                content = "\n".join(content_parts) if content_parts else ""

                child_node = {
                    "heading_text": heading_text if table_idx == 0 else "",
                    "heading_level": 2,
                    "content": content,
                    "children": [],
                }
                children.append(child_node)

        if not children and cell_data["data"]:
            all_content = []
            for row in cell_data["data"]:
                row_text = " | ".join(str(cell) for cell in row if cell)
                if row_text.strip():
                    all_content.append(row_text)

            if all_content:
                child_node = {
                    "heading_text": heading_text or "",
                    "heading_level": 2,
                    "content": "\n".join(all_content),
                    "children": [],
                }
                children.append(child_node)

        sheet_node["children"] = children
        return sheet_node

    def extract_xlsx_text(self, xlsx_path: str) -> List[Dict[str, Any]]:
        """
        Main entry point: Extract XLSX file and build document tree.

        Args:
            xlsx_path: Path to XLSX file

        Returns:
            List of sheet trees (array of sheets, each sheet is highest element)
        """
        self._xlsx_path = xlsx_path
        self._saved_images = []

        try:
            wb = load_workbook(xlsx_path, data_only=True)
            self._workbook = wb
        except Exception as exc:
            log.error("Failed to load workbook: %s", exc)
            return []

        # Step 1: Extract data from all sheets
        print('[DEBUG] EXTRACTION STEP')
        sheets_data = []
        for sheet_name in wb.sheetnames:
            print('[DEBUG] SHEET NAME', sheet_name)
            try:
                ws = wb[sheet_name]
                sheet_data = self._extract_sheet_data(ws, xlsx_path, sheet_name)
                sheets_data.append(sheet_data)
            except Exception as exc:
                log.warning("Error processing sheet %s: %s", sheet_name, exc)

        # Step 2: Build document tree for each sheet
        print('[DEBUG] BUILDING STEP')
        document_tree = []
        for sheet_data in sheets_data:
            print('[DEBUG] SHEET NAME', sheet_data["sheet_name"])
            try:
                sheet_tree = self._build_sheet_tree(sheet_data)
                document_tree.append(sheet_tree)
            except Exception as exc:
                log.warning("Error building tree for sheet: %s", exc)

        return document_tree

    def cleanup(self):
        """Clean up temporary image files."""
        if self._image_dir and os.path.exists(self._image_dir):
            try:
                import shutil
                shutil.rmtree(self._image_dir)
            except Exception as exc:
                log.warning("Failed to cleanup image dir: %s", exc)
        self._image_dir = None
        self._saved_images = []


# for unit test only
if __name__ == "__main__":
    import argparse

    arg_parser = argparse.ArgumentParser(description="Extract data from XLSX files")
    arg_parser.add_argument("--input", help="Input XLSX file path", default="SDF-QS-193-システムテスト項目書-20241019-20251211085426_sample_fromFPT.xlsx")
    arg_parser.add_argument(
        "--root-folder", "-rf", help="Root folder for working", default=""
    )
    arg_parser.add_argument(
        "--output", "-o", help="Output JSON file path", default="xlsx_output.json"
    )
    args = arg_parser.parse_args()

    parser = XlsxParser()
    result = parser.extract_xlsx_text(f"{args.root_folder}{args.input}")

    with open(f"{args.root_folder}{args.output}", "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"Output written to {args.output}")
