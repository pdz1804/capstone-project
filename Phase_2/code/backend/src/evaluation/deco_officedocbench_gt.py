"""Build OfficeDocBench-style ground truth from DECO spreadsheet annotations.

DECO annotations are required for strong GT. This module reads annotated table
regions and extracts the corresponding cell text from the workbook, producing
only fields consumed by the current OfficeDocBench metrics.
"""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from src.evaluation.officedocbench_adapter import empty_officedocbench_output
from src.evaluation.parsing_info_loss.utils import normalize_text

try:
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries
except Exception:  # pragma: no cover - surfaced as runtime error in converter
    load_workbook = None  # type: ignore
    range_boundaries = None  # type: ignore


@dataclass(frozen=True)
class TableRegion:
    sheet_name: str
    min_row: int
    min_col: int
    max_row: int
    max_col: int
    label: str = ""


def discover_deco_xlsx_jobs(source_dir: Path | str, annotation_root: Path | str | None = None) -> List[Dict[str, Any]]:
    root = Path(source_dir)
    annotations = Path(annotation_root) if annotation_root else None
    jobs: List[Dict[str, Any]] = []
    for source_path in sorted(root.glob("*")):
        if source_path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            continue
        annotation_path = _find_annotation(source_path, annotations) if annotations else None
        status = "OK" if annotation_path else "MISSING_ANNOTATION"
        if source_path.suffix.lower() == ".xls":
            status = "UNSUPPORTED_WORKBOOK"
        jobs.append(
            {
                "dataset": "deco-xlsx",
                "format": "xlsx",
                "file": source_path.name,
                "doc_id": source_path.stem,
                "source_path": str(source_path),
                "annotation_path": str(annotation_path) if annotation_path else None,
                "status": status,
            }
        )
    return jobs


def build_deco_officedocbench_gt(workbook_path: Path | str, annotation_path: Path | str) -> Dict[str, Any]:
    """Convert DECO table annotations into OfficeDocBench-style GT."""

    if load_workbook is None or range_boundaries is None:
        raise RuntimeError("openpyxl is required for DECO XLSX GT conversion")
    workbook_file = Path(workbook_path)
    if workbook_file.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(f"DECO GT conversion supports .xlsx/.xlsm workbooks, got: {workbook_file}")

    annotation_file = Path(annotation_path)
    wb = load_workbook(workbook_file, data_only=False, read_only=False)

    output = empty_officedocbench_output()
    output["metadata"] = {
        "sheet_names": list(wb.sheetnames),
    }

    if annotation_file.suffix.lower() == ".csv":
        regions = list(_extract_regions_from_range_csv(annotation_file, workbook_file.name))
    else:
        payload = json.loads(annotation_file.read_text(encoding="utf-8"))
        regions = list(_extract_regions(payload, default_sheet=wb.sheetnames[0] if wb.sheetnames else "Sheet1"))

    for region in regions:
        if region.sheet_name not in wb.sheetnames:
            continue
        ws = wb[region.sheet_name]
        rows = _rows_from_region(ws, region)
        if not rows:
            continue
        merge_count = _merge_count_for_region(ws, region)
        output["tables"].append(_table_payload(rows, merge_count, region))
        if region.label:
            output["text_elements"].append({"text": region.label, "style": "table_label"})

    return output


def write_deco_officedocbench_gt(workbook_path: Path | str, annotation_path: Path | str, output_path: Path | str) -> Dict[str, Any]:
    gt = build_deco_officedocbench_gt(workbook_path, annotation_path)
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(gt, ensure_ascii=False, indent=2), encoding="utf-8")
    return gt


def _find_annotation(source_path: Path, annotation_root: Optional[Path]) -> Optional[Path]:
    if annotation_root is None or not annotation_root.exists():
        return None
    if annotation_root.is_file() and annotation_root.suffix.lower() == ".csv":
        return annotation_root
    candidates = [
        annotation_root / f"{source_path.stem}.json",
        annotation_root / f"{source_path.name}.json",
        annotation_root / "range_annotations.csv",
        annotation_root / "rangeAnnotations.csv",
        annotation_root / source_path.stem / "annotation.json",
        annotation_root / source_path.stem / f"{source_path.stem}.json",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    matches = sorted(annotation_root.glob(f"**/{source_path.stem}.json"))
    return matches[0] if matches else None


def _extract_regions_from_range_csv(annotation_file: Path, file_name: str) -> Iterable[TableRegion]:
    with annotation_file.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if (row.get("FileName") or "").strip() != file_name:
                continue
            if (row.get("AnnotationLabel") or "").strip().lower() != "table":
                continue
            sheet_name = normalize_text(str(row.get("SheetName") or "Sheet1"))
            label = ""
            min_col = _coerce_int(row.get("FirstColumn"), -1) + 1
            min_row = _coerce_int(row.get("FirstRow"), -1) + 1
            max_col = _coerce_int(row.get("LastColumn"), -1) + 1
            max_row = _coerce_int(row.get("LastRow"), -1) + 1
            region = _validated_region(sheet_name, min_row, min_col, max_row, max_col, label)
            if region is not None:
                yield region


def _extract_regions(payload: Any, *, default_sheet: str) -> Iterable[TableRegion]:
    for item in _iter_annotation_items(payload):
        if not isinstance(item, dict):
            continue
        if not _looks_like_table_item(item):
            continue
        region = _region_from_item(item, default_sheet=default_sheet)
        if region is not None:
            yield region


def _iter_annotation_items(payload: Any) -> Iterable[Any]:
    if isinstance(payload, list):
        for item in payload:
            yield from _iter_annotation_items(item)
        return
    if not isinstance(payload, dict):
        return
    for key in ("tables", "table_regions", "regions", "annotations", "data"):
        value = payload.get(key)
        if isinstance(value, list):
            for item in value:
                yield from _iter_annotation_items(item)
            return
    yield payload


def _looks_like_table_item(item: Dict[str, Any]) -> bool:
    kind = normalize_text(str(item.get("type") or item.get("class") or item.get("category") or "")).lower()
    if kind and not any(token in kind for token in ("table", "region", "data")):
        return False
    return any(key in item for key in ("range", "ref", "bbox", "bounds", "min_row", "start_row", "top", "row_start"))


def _region_from_item(item: Dict[str, Any], *, default_sheet: str) -> Optional[TableRegion]:
    sheet_name = normalize_text(
        str(item.get("sheet_name") or item.get("sheet") or item.get("worksheet") or item.get("tab") or default_sheet)
    )
    label = normalize_text(str(item.get("table_label") or item.get("label_text") or item.get("name") or ""))

    range_value = item.get("range") or item.get("ref") or item.get("cell_range")
    if isinstance(range_value, str) and range_value.strip():
        try:
            min_col, min_row, max_col, max_row = range_boundaries(range_value.split("!", 1)[-1])
            return TableRegion(sheet_name, min_row, min_col, max_row, max_col, label)
        except Exception:
            pass

    bounds = item.get("bounds") or item.get("bbox")
    if isinstance(bounds, dict):
        item = {**item, **bounds}
    elif isinstance(bounds, list) and len(bounds) >= 4:
        min_row, min_col, max_row, max_col = [_coerce_int(v, 0) for v in bounds[:4]]
        return _validated_region(sheet_name, min_row, min_col, max_row, max_col, label)

    min_row = _first_int(item, ("min_row", "start_row", "row_start", "top", "first_row"))
    max_row = _first_int(item, ("max_row", "end_row", "row_end", "bottom", "last_row"))
    min_col = _first_int(item, ("min_col", "start_col", "col_start", "left", "first_col"))
    max_col = _first_int(item, ("max_col", "end_col", "col_end", "right", "last_col"))
    if None in (min_row, max_row, min_col, max_col):
        return None
    return _validated_region(sheet_name, min_row or 0, min_col or 0, max_row or 0, max_col or 0, label)


def _validated_region(sheet_name: str, min_row: int, min_col: int, max_row: int, max_col: int, label: str) -> Optional[TableRegion]:
    if min_row <= 0 or min_col <= 0 or max_row < min_row or max_col < min_col:
        return None
    return TableRegion(sheet_name, min_row, min_col, max_row, max_col, label)


def _first_int(item: Dict[str, Any], keys: Sequence[str]) -> Optional[int]:
    for key in keys:
        if key in item:
            return _coerce_int(item.get(key), 0)
    return None


def _rows_from_region(ws: Any, region: TableRegion) -> List[List[str]]:
    rows: List[List[str]] = []
    for row_idx in range(region.min_row, region.max_row + 1):
        row: List[str] = []
        for col_idx in range(region.min_col, region.max_col + 1):
            row.append(normalize_text(str(ws.cell(row_idx, col_idx).value or "")))
        if any(row):
            rows.append(row)
    return rows


def _merge_count_for_region(ws: Any, region: TableRegion) -> int:
    merge_count = 0
    for merge_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merge_range.bounds
        if max_row < region.min_row or min_row > region.max_row or max_col < region.min_col or min_col > region.max_col:
            continue
        overlap_rows = max(0, min(max_row, region.max_row) - max(min_row, region.min_row) + 1)
        overlap_cols = max(0, min(max_col, region.max_col) - max(min_col, region.min_col) + 1)
        if overlap_rows and overlap_cols:
            merge_count += max(overlap_rows * overlap_cols - 1, 1)
    return merge_count


def _table_payload(rows: Sequence[Sequence[str]], merge_count: int, region: TableRegion) -> Dict[str, Any]:
    normalized_rows = [[normalize_text(str(cell)) for cell in row] for row in rows if row]
    payload: Dict[str, Any] = {
        "rows": normalized_rows,
        "row_count": len(normalized_rows),
        "has_merged_cells": merge_count > 0,
        "cell_text": normalize_text(" ".join(" ".join(row) for row in normalized_rows)),
        "sheet_name": region.sheet_name,
        "range": {
            "min_row": region.min_row,
            "min_col": region.min_col,
            "max_row": region.max_row,
            "max_col": region.max_col,
        },
    }
    if merge_count:
        payload["merge_count"] = merge_count
    return payload


def _coerce_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except Exception:
        match = re.search(r"\d+", str(value or ""))
        return int(match.group(0)) if match else default
