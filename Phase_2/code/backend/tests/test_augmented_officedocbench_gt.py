from __future__ import annotations

import json

from openpyxl import Workbook

from src.evaluation.deco_officedocbench_gt import build_deco_officedocbench_gt
from src.evaluation.officedocbench_adapter import empty_officedocbench_output
from src.evaluation.pandoc_officedocbench_gt import build_pandoc_officedocbench_gt


def test_pandoc_native_maps_only_officedocbench_metric_fields(tmp_path):
    native_path = tmp_path / "case.native"
    native_path.write_text(
        """
        [ Header 2 ("intro",[],[]) [Str "Intro"]
        , Para [Str "Hello",Space,Link ("",[],[]) [Str "site"] ("https://example.com","")]
        , BulletList [[Plain [Str "One"]],[Plain [Str "Two"]]]
        , Div ("",["notes"],[]) [Para [Str "Speaker",Space,Str "note"]]
        , Para [Image ("",[],[]) [Str "Alt"] ("media/image.png","")]
        , Table
            ("",[],[])
            (Caption Nothing [])
            [(AlignDefault,0.5)]
            (TableHead ("",[],[]) [])
            [TableBody
              ("",[],[])
              (RowHeadColumns 0)
              []
              [Row ("",[],[])
                [Cell ("",[],[]) AlignDefault (RowSpan 1) (ColSpan 2) [Plain [Str "Merged"]]]
              ,Row ("",[],[])
                [Cell ("",[],[]) AlignDefault (RowSpan 1) (ColSpan 1) [Plain [Str "Value"]]]
              ]]
            ]
            (TableFoot ("",[],[]) [])
        ]
        """,
        encoding="utf-8",
    )
    source_path = tmp_path / "case.docx"
    source_path.write_bytes(b"not a real docx")

    gt = build_pandoc_officedocbench_gt(native_path, file_format="docx", source_path=source_path)

    assert set(gt) == set(empty_officedocbench_output())
    assert gt["headings"] == [{"text": "Intro", "level": 2}]
    assert gt["text_elements"][0]["text"] == "Hello site"
    assert gt["hyperlinks"] == [{"text": "site", "url": "https://example.com"}]
    assert gt["lists"] == [{"items": ["One", "Two"], "ordered": False, "depth": 1}]
    assert gt["speaker_notes"] == [{"text": "Speaker note"}]
    assert gt["images"] == [{"description": "Alt", "path": "media/image.png"}]
    assert gt["tables"][0]["rows"] == [["Merged"], ["Value"]]
    assert gt["tables"][0]["row_count"] == 2
    assert gt["tables"][0]["has_merged_cells"] is True
    assert gt["tables"][0]["merge_count"] == 1
    assert gt["comments"] == []
    assert gt["track_changes"] == []


def test_deco_annotation_maps_table_regions_and_sheet_metadata(tmp_path):
    workbook_path = tmp_path / "book.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue"
    ws["A1"] = "Region"
    ws["B1"] = "Q1"
    ws["A2"] = "North"
    ws["B2"] = 10
    ws.merge_cells("A4:B4")
    ws["A4"] = "Merged title"
    wb.save(workbook_path)

    annotation_path = tmp_path / "book.json"
    annotation_path.write_text(
        json.dumps(
            {
                "tables": [
                    {"sheet_name": "Revenue", "range": "A1:B2", "table_label": "Main table"},
                    {"sheet_name": "Revenue", "range": "A4:B4"},
                ]
            }
        ),
        encoding="utf-8",
    )

    gt = build_deco_officedocbench_gt(workbook_path, annotation_path)

    assert set(gt) == set(empty_officedocbench_output())
    assert gt["metadata"]["sheet_names"] == ["Revenue"]
    assert gt["text_elements"] == [{"text": "Main table", "style": "table_label"}]
    assert gt["tables"][0]["rows"] == [["Region", "Q1"], ["North", "10"]]
    assert gt["tables"][0]["row_count"] == 2
    assert gt["tables"][0]["has_merged_cells"] is False
    assert gt["tables"][1]["rows"] == [["Merged title", ""]]
    assert gt["tables"][1]["has_merged_cells"] is True
    assert gt["tables"][1]["merge_count"] == 1
    assert gt["headings"] == []
    assert gt["lists"] == []

