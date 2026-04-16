from __future__ import annotations

from argparse import ArgumentParser
from dataclasses import dataclass
from datetime import date
from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="1F4E78")
SUBTITLE_FONT = Font(size=10, italic=True, color="666666")
VALUE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

SAGEMAKER_INSTANCE_PRICE_HINTS = {
    "ml.g4dn.xlarge": 0.7364,
    "ml.g5.xlarge": 1.4080,
    "ml.m5.large": 0.1340,
    "ml.c5.xlarge": 0.2040,
}


@dataclass
class TerraformContext:
    terraform_dir: Path
    variables_source: str
    resource_types: set[str]
    region: str
    project_name: str
    environment: str
    backend_desired_count: float
    frontend_desired_count: float
    backend_task_vcpu: float
    backend_task_memory_gb: float
    frontend_task_vcpu: float
    frontend_task_memory_gb: float
    enable_sagemaker_endpoint: bool
    sagemaker_instance_type: str
    sagemaker_initial_instance_count: float
    enable_waf: bool
    enable_search_cache_serverless: bool
    enable_chatbot_history_tables: bool
    enable_agentcore_runtime_prep: bool
    ecr_repository_count: int


@dataclass
class CostComponent:
    key: str
    service: str
    item: str
    unit: str
    unit_price: float
    usage_quantity: float
    terraform_basis: str
    pricing_url: str
    notes: str


def month_sequence(start_month: date, count: int) -> list[date]:
    months: list[date] = []
    year = start_month.year
    month = start_month.month

    for _ in range(count):
        months.append(date(year, month, 1))
        month += 1
        if month > 12:
            month = 1
            year += 1

    return months


def style_header_row(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def apply_borders(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            cell.border = THIN_BORDER


def set_column_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def strip_inline_comment(raw: str) -> str:
    result: list[str] = []
    in_quotes = False
    escaped = False
    i = 0

    while i < len(raw):
        ch = raw[i]
        next_ch = raw[i + 1] if i + 1 < len(raw) else ""

        if ch == '"' and not escaped:
            in_quotes = not in_quotes
            result.append(ch)
        elif not in_quotes and ch == "#":
            break
        elif not in_quotes and ch == "/" and next_ch == "/":
            break
        else:
            result.append(ch)

        if ch == "\\" and not escaped:
            escaped = True
        else:
            escaped = False

        i += 1

    return "".join(result).strip()


def parse_hcl_literal(raw: str) -> Any:
    text = strip_inline_comment(raw).strip()
    if text == "":
        return ""

    if text.startswith('"') and text.endswith('"') and len(text) >= 2:
        return text[1:-1].replace('\\"', '"')

    lowered = text.lower()
    if lowered == "true":
        return True
    if lowered == "false":
        return False

    if text == "[]":
        return []
    if text == "{}":
        return {}

    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d+\.\d+", text):
        return float(text)

    return text


def parse_variable_defaults(variables_file: Path) -> dict[str, Any]:
    defaults: dict[str, Any] = {}
    if not variables_file.exists():
        return defaults

    lines = variables_file.read_text(encoding="utf-8").splitlines()
    current_var: str | None = None
    block_lines: list[str] = []
    brace_depth = 0

    for line in lines:
        if current_var is None:
            match = re.match(r"\s*variable\s+\"([^\"]+)\"\s*{", line)
            if match:
                current_var = match.group(1)
                block_lines = [line]
                brace_depth = line.count("{") - line.count("}")
                if brace_depth == 0:
                    value = extract_default_from_variable_block(block_lines)
                    if value is not None:
                        defaults[current_var] = value
                    current_var = None
                    block_lines = []
        else:
            block_lines.append(line)
            brace_depth += line.count("{") - line.count("}")
            if brace_depth == 0:
                value = extract_default_from_variable_block(block_lines)
                if value is not None:
                    defaults[current_var] = value
                current_var = None
                block_lines = []

    return defaults


def extract_default_from_variable_block(block_lines: list[str]) -> Any | None:
    for line in block_lines:
        match = re.match(r"\s*default\s*=\s*(.+)$", line)
        if not match:
            continue

        value_raw = match.group(1).strip()
        if value_raw == "{":
            return {}
        if value_raw == "[":
            return []
        return parse_hcl_literal(value_raw)

    return None


def parse_tfvars_file(tfvars_file: Path) -> dict[str, Any]:
    values: dict[str, Any] = {}
    if not tfvars_file.exists():
        return values

    for raw_line in tfvars_file.read_text(encoding="utf-8").splitlines():
        stripped = raw_line.strip()
        if stripped == "" or stripped.startswith("#") or stripped.startswith("//"):
            continue
        if "=" not in stripped:
            continue

        key, value_raw = stripped.split("=", 1)
        key = key.strip()
        value = parse_hcl_literal(value_raw)
        values[key] = value

    return values


def read_terraform_texts(terraform_dir: Path) -> dict[Path, str]:
    texts: dict[Path, str] = {}
    for tf_path in sorted(terraform_dir.rglob("*.tf")):
        texts[tf_path] = tf_path.read_text(encoding="utf-8")
    return texts


def extract_resource_types(terraform_texts: dict[Path, str]) -> set[str]:
    resource_types: set[str] = set()
    pattern = re.compile(r'resource\s+"(aws_[^\"]+)"\s+"[^\"]+"')

    for text in terraform_texts.values():
        for match in pattern.finditer(text):
            resource_types.add(match.group(1))

    return resource_types


def extract_object_block(text: str, object_name: str) -> str:
    match = re.search(rf"\b{re.escape(object_name)}\s*=\s*{{", text)
    if not match:
        return ""

    start_idx = text.find("{", match.start())
    if start_idx < 0:
        return ""

    depth = 0
    for idx in range(start_idx, len(text)):
        char = text[idx]
        if char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                return text[start_idx : idx + 1]

    return ""


def extract_numeric_setting_from_object(text: str, object_name: str, key: str, default: float) -> float:
    block = extract_object_block(text, object_name)
    if block == "":
        return default

    match = re.search(rf"\b{re.escape(key)}\s*=\s*([0-9]+(?:\.[0-9]+)?)", block)
    if not match:
        return default

    return float(match.group(1))


def as_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "yes", "1"}:
            return True
        if lowered in {"false", "no", "0"}:
            return False
    if isinstance(value, (int, float)):
        return bool(value)
    return default


def as_float(value: Any, default: float = 0.0) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
            return float(text)
    return default


def discover_terraform_context(terraform_dir: Path) -> TerraformContext:
    variables_file = terraform_dir / "variables.tf"
    main_file = terraform_dir / "main.tf"

    defaults = parse_variable_defaults(variables_file)

    vars_candidates = [
        terraform_dir / "terraform.tfvars",
        terraform_dir / "terraform.auto.tfvars",
        terraform_dir / "terraform.tfvars.example",
    ]

    merged_values = defaults.copy()
    vars_source = "variables.tf defaults"

    for candidate in vars_candidates:
        if candidate.exists():
            merged_values.update(parse_tfvars_file(candidate))
            vars_source = candidate.name
            break

    terraform_texts = read_terraform_texts(terraform_dir)
    resource_types = extract_resource_types(terraform_texts)

    main_text = main_file.read_text(encoding="utf-8") if main_file.exists() else ""

    backend_cpu_units = extract_numeric_setting_from_object(main_text, "backend_config", "cpu", 512.0)
    backend_memory_mib = extract_numeric_setting_from_object(main_text, "backend_config", "memory", 1024.0)
    frontend_cpu_units = extract_numeric_setting_from_object(main_text, "frontend_config", "cpu", 256.0)
    frontend_memory_mib = extract_numeric_setting_from_object(main_text, "frontend_config", "memory", 512.0)

    enable_sagemaker_endpoint = as_bool(merged_values.get("enable_sagemaker_endpoint", True), True)
    enable_waf = as_bool(merged_values.get("enable_waf", False), False)
    enable_search_cache_serverless = as_bool(merged_values.get("enable_search_cache_serverless", True), True)
    enable_chatbot_history_tables = as_bool(merged_values.get("enable_chatbot_history_tables", True), True)
    enable_agentcore_runtime_prep = as_bool(merged_values.get("enable_agentcore_runtime_prep", True), True)

    ecr_repository_count = 2 + int(enable_agentcore_runtime_prep) + int(enable_sagemaker_endpoint)

    return TerraformContext(
        terraform_dir=terraform_dir,
        variables_source=vars_source,
        resource_types=resource_types,
        region=str(merged_values.get("aws_region", "us-west-2")),
        project_name=str(merged_values.get("project_name", "rag-pipeline")),
        environment=str(merged_values.get("environment", "prod")),
        backend_desired_count=as_float(merged_values.get("backend_desired_count", 2), 2),
        frontend_desired_count=as_float(merged_values.get("frontend_desired_count", 2), 2),
        backend_task_vcpu=backend_cpu_units / 1024.0,
        backend_task_memory_gb=backend_memory_mib / 1024.0,
        frontend_task_vcpu=frontend_cpu_units / 1024.0,
        frontend_task_memory_gb=frontend_memory_mib / 1024.0,
        enable_sagemaker_endpoint=enable_sagemaker_endpoint,
        sagemaker_instance_type=str(merged_values.get("sagemaker_instance_type", "ml.g4dn.xlarge")),
        sagemaker_initial_instance_count=as_float(merged_values.get("sagemaker_initial_instance_count", 1), 1),
        enable_waf=enable_waf,
        enable_search_cache_serverless=enable_search_cache_serverless,
        enable_chatbot_history_tables=enable_chatbot_history_tables,
        enable_agentcore_runtime_prep=enable_agentcore_runtime_prep,
        ecr_repository_count=ecr_repository_count,
    )


def build_assumption_rows(context: TerraformContext) -> list[tuple[str, Any, str, str, str, bool]]:
    detected_resource_types = ", ".join(sorted(context.resource_types))
    sagemaker_avg_instances = context.sagemaker_initial_instance_count if context.enable_sagemaker_endpoint else 0.0

    return [
        (
            "terraform_directory",
            str(context.terraform_dir),
            "text",
            "filesystem",
            "Terraform folder used for resource discovery.",
            False,
        ),
        (
            "terraform_variables_source",
            context.variables_source,
            "text",
            "filesystem",
            "Variable source merged on top of defaults.",
            False,
        ),
        (
            "detected_aws_resource_types",
            detected_resource_types,
            "text",
            "terraform scan",
            "Detected aws_* resource types that drive service mapping.",
            False,
        ),
        (
            "aws_region",
            context.region,
            "region",
            "terraform variable",
            "Primary region used for price references.",
            False,
        ),
        (
            "project_name",
            context.project_name,
            "text",
            "terraform variable",
            "Project naming root.",
            False,
        ),
        (
            "environment",
            context.environment,
            "text",
            "terraform variable",
            "Terraform environment value.",
            False,
        ),
        (
            "backend_desired_count",
            context.backend_desired_count,
            "count",
            "terraform variable",
            "ECS backend desired task count.",
            False,
        ),
        (
            "frontend_desired_count",
            context.frontend_desired_count,
            "count",
            "terraform variable",
            "ECS frontend desired task count.",
            False,
        ),
        (
            "backend_task_vcpu",
            context.backend_task_vcpu,
            "vCPU",
            "terraform module",
            "Derived from modules/ecs backend_config.cpu / 1024.",
            False,
        ),
        (
            "backend_task_memory_gb",
            context.backend_task_memory_gb,
            "GB",
            "terraform module",
            "Derived from modules/ecs backend_config.memory / 1024.",
            False,
        ),
        (
            "frontend_task_vcpu",
            context.frontend_task_vcpu,
            "vCPU",
            "terraform module",
            "Derived from modules/ecs frontend_config.cpu / 1024.",
            False,
        ),
        (
            "frontend_task_memory_gb",
            context.frontend_task_memory_gb,
            "GB",
            "terraform module",
            "Derived from modules/ecs frontend_config.memory / 1024.",
            False,
        ),
        (
            "ecr_repository_count",
            context.ecr_repository_count,
            "count",
            "terraform modules",
            "backend + frontend + optional agentcore + optional sagemaker repo.",
            False,
        ),
        (
            "enable_sagemaker_endpoint",
            context.enable_sagemaker_endpoint,
            "boolean",
            "terraform variable",
            "Controls SageMaker endpoint cost rows.",
            False,
        ),
        (
            "sagemaker_instance_type",
            context.sagemaker_instance_type,
            "text",
            "terraform variable",
            "SageMaker production variant instance type.",
            False,
        ),
        (
            "enable_waf",
            context.enable_waf,
            "boolean",
            "terraform variable",
            "Controls WAF pricing rows.",
            False,
        ),
        (
            "enable_search_cache_serverless",
            context.enable_search_cache_serverless,
            "boolean",
            "terraform variable",
            "Controls ElastiCache Serverless pricing rows.",
            False,
        ),
        (
            "enable_chatbot_history_tables",
            context.enable_chatbot_history_tables,
            "boolean",
            "terraform variable",
            "Controls DynamoDB pricing rows.",
            False,
        ),
        (
            "monthly_hours",
            730.0,
            "hours",
            "estimation assumption",
            "Average billable hours in one month.",
            True,
        ),
        (
            "contingency_buffer_pct",
            0.15,
            "percent",
            "estimation assumption",
            "Risk buffer applied to base monthly cost.",
            True,
        ),
        (
            "avg_alb_lcu",
            1.0,
            "LCU",
            "estimation assumption",
            "Average ALB LCU utilization.",
            True,
        ),
        (
            "cloudwatch_log_ingest_gb",
            60.0,
            "GB",
            "estimation assumption",
            "Monthly log ingestion across ECS and WAF logs.",
            True,
        ),
        (
            "cloudwatch_log_storage_gb",
            30.0,
            "GB",
            "estimation assumption",
            "Average retained log volume billed per month.",
            True,
        ),
        (
            "ecr_storage_gb_per_repo",
            4.0,
            "GB",
            "estimation assumption",
            "Average image storage per ECR repository.",
            True,
        ),
        (
            "dynamodb_read_million",
            40.0 if context.enable_chatbot_history_tables else 0.0,
            "million requests",
            "estimation assumption",
            "On-demand strongly consistent reads (million units/month).",
            True,
        ),
        (
            "dynamodb_write_million",
            12.0 if context.enable_chatbot_history_tables else 0.0,
            "million requests",
            "estimation assumption",
            "On-demand write request units (million units/month).",
            True,
        ),
        (
            "dynamodb_storage_gb",
            8.0 if context.enable_chatbot_history_tables else 0.0,
            "GB",
            "estimation assumption",
            "Average DynamoDB storage for managed chatbot tables.",
            True,
        ),
        (
            "search_cache_avg_data_gb",
            2.0 if context.enable_search_cache_serverless else 0.0,
            "GB",
            "estimation assumption",
            "Average data stored in ElastiCache Serverless.",
            True,
        ),
        (
            "search_cache_ecpu_million",
            300.0 if context.enable_search_cache_serverless else 0.0,
            "million ECPU",
            "estimation assumption",
            "ElastiCache serverless compute usage in million ECPU.",
            True,
        ),
        (
            "sagemaker_avg_instance_count",
            sagemaker_avg_instances,
            "count",
            "estimation assumption",
            "Average running SageMaker instances over a month.",
            True,
        ),
        (
            "waf_requests_million",
            80.0 if context.enable_waf else 0.0,
            "million requests",
            "estimation assumption",
            "WAF inspected requests billed per month.",
            True,
        ),
        (
            "price_reference_last_updated",
            date.today(),
            "date",
            "generated",
            "Pricing figures are reference values and must be verified.",
            False,
        ),
    ]


def build_assumptions_sheet(ws_assumptions, assumption_rows: list[tuple[str, Any, str, str, str, bool]]) -> dict[str, int]:
    headers = ["Parameter Key", "Value", "Unit", "Source", "Notes"]
    style_header_row(ws_assumptions, headers)

    row_lookup: dict[str, int] = {}

    for row_idx, (key, value, unit, source, notes, editable) in enumerate(assumption_rows, start=2):
        ws_assumptions.cell(row=row_idx, column=1, value=key)
        value_cell = ws_assumptions.cell(row=row_idx, column=2, value=value)
        ws_assumptions.cell(row=row_idx, column=3, value=unit)
        ws_assumptions.cell(row=row_idx, column=4, value=source)
        ws_assumptions.cell(row=row_idx, column=5, value=notes)

        if editable:
            value_cell.fill = VALUE_FILL
            value_cell.font = Font(bold=True)

        if unit == "percent":
            value_cell.number_format = "0.00%"
        elif unit in {"count", "hours", "vCPU", "GB", "LCU", "million requests", "million ECPU"}:
            value_cell.number_format = "#,##0.00"
        elif unit == "date":
            value_cell.number_format = "yyyy-mm-dd"

        row_lookup[key] = row_idx

    set_column_widths(
        ws_assumptions,
        {
            "A": 36,
            "B": 22,
            "C": 18,
            "D": 22,
            "E": 72,
        },
    )
    ws_assumptions.freeze_panes = "A2"
    apply_borders(ws_assumptions, start_row=1, end_row=1 + len(assumption_rows), start_col=1, end_col=5)

    return row_lookup


def assumption_float(assumption_values: dict[str, Any], key: str, default: float = 0.0) -> float:
    return as_float(assumption_values.get(key, default), default)


def build_cost_components(context: TerraformContext, assumption_values: dict[str, Any]) -> list[CostComponent]:
    components: list[CostComponent] = []

    monthly_hours = assumption_float(assumption_values, "monthly_hours", 730.0)
    alb_lcu = assumption_float(assumption_values, "avg_alb_lcu", 1.0)
    log_ingest_gb = assumption_float(assumption_values, "cloudwatch_log_ingest_gb", 60.0)
    log_storage_gb = assumption_float(assumption_values, "cloudwatch_log_storage_gb", 30.0)
    ecr_storage_per_repo = assumption_float(assumption_values, "ecr_storage_gb_per_repo", 4.0)
    ddb_read_million = assumption_float(assumption_values, "dynamodb_read_million", 0.0)
    ddb_write_million = assumption_float(assumption_values, "dynamodb_write_million", 0.0)
    ddb_storage_gb = assumption_float(assumption_values, "dynamodb_storage_gb", 0.0)
    cache_avg_data_gb = assumption_float(assumption_values, "search_cache_avg_data_gb", 0.0)
    cache_ecpu_million = assumption_float(assumption_values, "search_cache_ecpu_million", 0.0)
    sagemaker_avg_instances = assumption_float(assumption_values, "sagemaker_avg_instance_count", 0.0)
    waf_requests_million = assumption_float(assumption_values, "waf_requests_million", 0.0)

    resource_types = context.resource_types

    if "aws_ecs_service" in resource_types:
        backend_vcpu_hours = monthly_hours * context.backend_desired_count * context.backend_task_vcpu
        backend_memory_gb_hours = monthly_hours * context.backend_desired_count * context.backend_task_memory_gb
        frontend_vcpu_hours = monthly_hours * context.frontend_desired_count * context.frontend_task_vcpu
        frontend_memory_gb_hours = monthly_hours * context.frontend_desired_count * context.frontend_task_memory_gb

        components.extend(
            [
                CostComponent(
                    key="ECS_Fargate_Backend_vCPU_Hr",
                    service="Amazon ECS (Fargate)",
                    item="Backend Fargate vCPU usage",
                    unit="vCPU-hour",
                    unit_price=0.04048,
                    usage_quantity=backend_vcpu_hours,
                    terraform_basis="aws_ecs_service.backend + aws_ecs_task_definition.backend",
                    pricing_url="https://aws.amazon.com/fargate/pricing/",
                    notes="Derived from backend desired_count and backend task vCPU from Terraform.",
                ),
                CostComponent(
                    key="ECS_Fargate_Backend_GB_Hr",
                    service="Amazon ECS (Fargate)",
                    item="Backend Fargate memory usage",
                    unit="GB-hour",
                    unit_price=0.004445,
                    usage_quantity=backend_memory_gb_hours,
                    terraform_basis="aws_ecs_service.backend + aws_ecs_task_definition.backend",
                    pricing_url="https://aws.amazon.com/fargate/pricing/",
                    notes="Derived from backend desired_count and backend task memory from Terraform.",
                ),
                CostComponent(
                    key="ECS_Fargate_Frontend_vCPU_Hr",
                    service="Amazon ECS (Fargate)",
                    item="Frontend Fargate vCPU usage",
                    unit="vCPU-hour",
                    unit_price=0.04048,
                    usage_quantity=frontend_vcpu_hours,
                    terraform_basis="aws_ecs_service.frontend + aws_ecs_task_definition.frontend",
                    pricing_url="https://aws.amazon.com/fargate/pricing/",
                    notes="Derived from frontend desired_count and frontend task vCPU from Terraform.",
                ),
                CostComponent(
                    key="ECS_Fargate_Frontend_GB_Hr",
                    service="Amazon ECS (Fargate)",
                    item="Frontend Fargate memory usage",
                    unit="GB-hour",
                    unit_price=0.004445,
                    usage_quantity=frontend_memory_gb_hours,
                    terraform_basis="aws_ecs_service.frontend + aws_ecs_task_definition.frontend",
                    pricing_url="https://aws.amazon.com/fargate/pricing/",
                    notes="Derived from frontend desired_count and frontend task memory from Terraform.",
                ),
            ]
        )

    if "aws_lb" in resource_types:
        components.extend(
            [
                CostComponent(
                    key="ALB_Hours",
                    service="Elastic Load Balancing (ALB)",
                    item="Application Load Balancer running hours",
                    unit="ALB-hour",
                    unit_price=0.0225,
                    usage_quantity=monthly_hours,
                    terraform_basis="aws_lb.main",
                    pricing_url="https://aws.amazon.com/elasticloadbalancing/pricing/",
                    notes="One ALB declared in modules/alb.",
                ),
                CostComponent(
                    key="ALB_LCU_Hours",
                    service="Elastic Load Balancing (ALB)",
                    item="ALB LCU usage",
                    unit="LCU-hour",
                    unit_price=0.0080,
                    usage_quantity=monthly_hours * alb_lcu,
                    terraform_basis="aws_lb.main + aws_lb_listener* + aws_lb_target_group*",
                    pricing_url="https://aws.amazon.com/elasticloadbalancing/pricing/",
                    notes="Average LCU is controlled by assumptions.avg_alb_lcu.",
                ),
            ]
        )

    if "aws_cloudwatch_log_group" in resource_types:
        components.extend(
            [
                CostComponent(
                    key="CloudWatch_Logs_Ingest_GB",
                    service="Amazon CloudWatch Logs",
                    item="Logs ingestion",
                    unit="GB",
                    unit_price=0.50,
                    usage_quantity=log_ingest_gb,
                    terraform_basis="aws_cloudwatch_log_group.backend/frontend and optional waf logs",
                    pricing_url="https://aws.amazon.com/cloudwatch/pricing/",
                    notes="Adjust ingest volume in assumptions.",
                ),
                CostComponent(
                    key="CloudWatch_Logs_Storage_GB_Month",
                    service="Amazon CloudWatch Logs",
                    item="Logs archival storage",
                    unit="GB",
                    unit_price=0.03,
                    usage_quantity=log_storage_gb,
                    terraform_basis="aws_cloudwatch_log_group.*",
                    pricing_url="https://aws.amazon.com/cloudwatch/pricing/",
                    notes="Adjust retained log volume in assumptions.",
                ),
            ]
        )

    if "aws_ecr_repository" in resource_types:
        components.append(
            CostComponent(
                key="ECR_Storage_GB_Month",
                service="Amazon ECR",
                item="Container image storage",
                unit="GB",
                unit_price=0.10,
                usage_quantity=context.ecr_repository_count * ecr_storage_per_repo,
                terraform_basis="module backend_ecr/fronted_ecr and optional AgentCore/SageMaker repos",
                pricing_url="https://aws.amazon.com/ecr/pricing/",
                notes="Storage estimate = repository_count * assumptions.ecr_storage_gb_per_repo.",
            )
        )

    if "aws_dynamodb_table" in resource_types and context.enable_chatbot_history_tables:
        components.extend(
            [
                CostComponent(
                    key="DynamoDB_OnDemand_Read_Million",
                    service="Amazon DynamoDB",
                    item="On-demand read request units",
                    unit="million requests",
                    unit_price=0.125,
                    usage_quantity=ddb_read_million,
                    terraform_basis="aws_dynamodb_table.chatbot_sessions/messages",
                    pricing_url="https://aws.amazon.com/dynamodb/pricing/on-demand/",
                    notes="Adjust read volume in assumptions.",
                ),
                CostComponent(
                    key="DynamoDB_OnDemand_Write_Million",
                    service="Amazon DynamoDB",
                    item="On-demand write request units",
                    unit="million requests",
                    unit_price=0.625,
                    usage_quantity=ddb_write_million,
                    terraform_basis="aws_dynamodb_table.chatbot_sessions/messages",
                    pricing_url="https://aws.amazon.com/dynamodb/pricing/on-demand/",
                    notes="Adjust write volume in assumptions.",
                ),
                CostComponent(
                    key="DynamoDB_Storage_GB_Month",
                    service="Amazon DynamoDB",
                    item="Data storage",
                    unit="GB",
                    unit_price=0.25,
                    usage_quantity=ddb_storage_gb,
                    terraform_basis="aws_dynamodb_table.chatbot_sessions/messages",
                    pricing_url="https://aws.amazon.com/dynamodb/pricing/on-demand/",
                    notes="Adjust stored data size in assumptions.",
                ),
            ]
        )

    if "aws_elasticache_serverless_cache" in resource_types and context.enable_search_cache_serverless:
        components.extend(
            [
                CostComponent(
                    key="ElastiCache_Serverless_Data_GB_Hour",
                    service="Amazon ElastiCache Serverless",
                    item="Data storage usage",
                    unit="GB-hour",
                    unit_price=0.125,
                    usage_quantity=cache_avg_data_gb * monthly_hours,
                    terraform_basis="aws_elasticache_serverless_cache.search_cache",
                    pricing_url="https://aws.amazon.com/elasticache/pricing/",
                    notes="Estimate uses average data GB from assumptions.",
                ),
                CostComponent(
                    key="ElastiCache_Serverless_ECPU_Million",
                    service="Amazon ElastiCache Serverless",
                    item="ECPU compute usage",
                    unit="million ECPU",
                    unit_price=0.20,
                    usage_quantity=cache_ecpu_million,
                    terraform_basis="aws_elasticache_serverless_cache.search_cache",
                    pricing_url="https://aws.amazon.com/elasticache/pricing/",
                    notes="Estimate uses million ECPU assumption.",
                ),
            ]
        )

    if "aws_sagemaker_endpoint" in resource_types and context.enable_sagemaker_endpoint:
        sagemaker_price = SAGEMAKER_INSTANCE_PRICE_HINTS.get(context.sagemaker_instance_type, 0.25)
        key_instance = re.sub(r"[^A-Za-z0-9]+", "_", context.sagemaker_instance_type)

        components.append(
            CostComponent(
                key=f"SageMaker_{key_instance}_Instance_Hour",
                service="Amazon SageMaker",
                item=f"Real-time endpoint hosting ({context.sagemaker_instance_type})",
                unit="instance-hour",
                unit_price=sagemaker_price,
                usage_quantity=sagemaker_avg_instances * monthly_hours,
                terraform_basis="aws_sagemaker_endpoint.main + aws_sagemaker_endpoint_configuration.main",
                pricing_url="https://aws.amazon.com/sagemaker/pricing/",
                notes="Average running instances comes from assumptions.sagemaker_avg_instance_count.",
            )
        )

    if "aws_wafv2_web_acl" in resource_types and context.enable_waf:
        components.extend(
            [
                CostComponent(
                    key="WAF_WebACL_Month",
                    service="AWS WAF",
                    item="Web ACL fixed monthly charge",
                    unit="month",
                    unit_price=5.00,
                    usage_quantity=1.0,
                    terraform_basis="aws_wafv2_web_acl.this",
                    pricing_url="https://aws.amazon.com/waf/pricing/",
                    notes="One Web ACL associated with ALB in modules/waf.",
                ),
                CostComponent(
                    key="WAF_Requests_Million",
                    service="AWS WAF",
                    item="Inspected requests",
                    unit="million requests",
                    unit_price=0.60,
                    usage_quantity=waf_requests_million,
                    terraform_basis="aws_wafv2_web_acl.this + aws_wafv2_web_acl_association.alb",
                    pricing_url="https://aws.amazon.com/waf/pricing/",
                    notes="Request volume estimate from assumptions.waf_requests_million.",
                ),
            ]
        )

    if not components:
        components.append(
            CostComponent(
                key="NoBillableResourcesDetected",
                service="Review Terraform",
                item="No billable aws_* resources were detected",
                unit="count",
                unit_price=0.0,
                usage_quantity=0.0,
                terraform_basis="resource scan result empty",
                pricing_url="https://aws.amazon.com/pricing/",
                notes="Confirm terraform path and aws resource declarations.",
            )
        )

    return components


def build_price_reference_sheet(ws_price, context: TerraformContext, components: list[CostComponent]) -> None:
    headers = [
        "Cost Component Key",
        "AWS Service",
        "Item",
        "Unit",
        "Price (USD)",
        "Region",
        "Pricing URL",
        "Terraform Basis",
        "Last Updated",
        "Notes",
    ]
    style_header_row(ws_price, headers)

    today = date.today()

    for row_idx, component in enumerate(components, start=2):
        ws_price.cell(row=row_idx, column=1, value=component.key)
        ws_price.cell(row=row_idx, column=2, value=component.service)
        ws_price.cell(row=row_idx, column=3, value=component.item)
        ws_price.cell(row=row_idx, column=4, value=component.unit)
        ws_price.cell(row=row_idx, column=5, value=component.unit_price)
        ws_price.cell(row=row_idx, column=6, value=context.region)
        ws_price.cell(row=row_idx, column=7, value=component.pricing_url)
        ws_price.cell(row=row_idx, column=8, value=component.terraform_basis)
        ws_price.cell(row=row_idx, column=9, value=today)
        ws_price.cell(row=row_idx, column=10, value=component.notes)

        ws_price.cell(row=row_idx, column=5).number_format = "$#,##0.0000"
        ws_price.cell(row=row_idx, column=9).number_format = "yyyy-mm-dd"

    set_column_widths(
        ws_price,
        {
            "A": 38,
            "B": 24,
            "C": 40,
            "D": 20,
            "E": 14,
            "F": 16,
            "G": 44,
            "H": 46,
            "I": 14,
            "J": 48,
        },
    )

    ws_price.freeze_panes = "A2"
    ws_price.auto_filter.ref = f"A1:J{1 + len(components)}"
    apply_borders(ws_price, start_row=1, end_row=1 + len(components), start_col=1, end_col=10)


def build_monthly_cost_sheet(
    ws_monthly,
    months: list[date],
    components: list[CostComponent],
    assumptions_row_lookup: dict[str, int],
) -> tuple[list[str], int]:
    headers = [
        "Month",
        "Cost Component Key",
        "AWS Service",
        "Usage Quantity",
        "Usage Unit",
        "Unit Price (USD)",
        "Base Cost (USD)",
        "Buffer (%)",
        "Total Cost (USD)",
        "Terraform Basis",
        "Notes",
    ]
    style_header_row(ws_monthly, headers)

    row_idx = 2
    services_seen: list[str] = []
    buffer_row = assumptions_row_lookup["contingency_buffer_pct"]

    for month in months:
        for component in components:
            ws_monthly.cell(row=row_idx, column=1, value=month)
            ws_monthly.cell(row=row_idx, column=2, value=component.key)
            ws_monthly.cell(row=row_idx, column=3, value=component.service)
            ws_monthly.cell(row=row_idx, column=4, value=component.usage_quantity)
            ws_monthly.cell(row=row_idx, column=5, value=component.unit)
            ws_monthly.cell(
                row=row_idx,
                column=6,
                value=(
                    f"=IFERROR(INDEX('aws price reference'!$E:$E,"
                    f"MATCH(B{row_idx},'aws price reference'!$A:$A,0)),0)"
                ),
            )
            ws_monthly.cell(row=row_idx, column=7, value=f"=D{row_idx}*F{row_idx}")
            ws_monthly.cell(row=row_idx, column=8, value=f"=IFERROR('assumptions'!$B${buffer_row},0)")
            ws_monthly.cell(row=row_idx, column=9, value=f"=G{row_idx}*(1+H{row_idx})")
            ws_monthly.cell(row=row_idx, column=10, value=component.terraform_basis)
            ws_monthly.cell(row=row_idx, column=11, value=component.notes)

            ws_monthly.cell(row=row_idx, column=1).number_format = "yyyy-mm"
            ws_monthly.cell(row=row_idx, column=4).number_format = "#,##0.00"
            ws_monthly.cell(row=row_idx, column=6).number_format = "$#,##0.0000"
            ws_monthly.cell(row=row_idx, column=7).number_format = "$#,##0.00"
            ws_monthly.cell(row=row_idx, column=8).number_format = "0.00%"
            ws_monthly.cell(row=row_idx, column=9).number_format = "$#,##0.00"

            if component.service not in services_seen:
                services_seen.append(component.service)

            row_idx += 1

    last_data_row = row_idx - 1

    ws_monthly.cell(row=row_idx + 1, column=6, value="TOTAL")
    ws_monthly.cell(row=row_idx + 1, column=7, value=f"=SUM(G2:G{last_data_row})")
    ws_monthly.cell(row=row_idx + 1, column=9, value=f"=SUM(I2:I{last_data_row})")
    ws_monthly.cell(row=row_idx + 1, column=6).font = Font(bold=True)
    ws_monthly.cell(row=row_idx + 1, column=7).font = Font(bold=True)
    ws_monthly.cell(row=row_idx + 1, column=9).font = Font(bold=True)
    ws_monthly.cell(row=row_idx + 1, column=7).number_format = "$#,##0.00"
    ws_monthly.cell(row=row_idx + 1, column=9).number_format = "$#,##0.00"

    set_column_widths(
        ws_monthly,
        {
            "A": 12,
            "B": 38,
            "C": 24,
            "D": 16,
            "E": 20,
            "F": 16,
            "G": 16,
            "H": 12,
            "I": 16,
            "J": 46,
            "K": 48,
        },
    )

    ws_monthly.freeze_panes = "A2"
    ws_monthly.auto_filter.ref = f"A1:K{last_data_row}"
    apply_borders(ws_monthly, start_row=1, end_row=row_idx + 1, start_col=1, end_col=11)

    return services_seen, last_data_row


def build_summary_dashboard(
    ws_summary,
    months: list[date],
    services: list[str],
    monthly_last_data_row: int,
    terraform_context: TerraformContext,
) -> None:
    ws_summary["A1"] = "Summary Dashboard"
    ws_summary["A1"].font = TITLE_FONT
    ws_summary["A2"] = (
        f"Terraform-aligned estimate for {terraform_context.project_name} in {terraform_context.region}."
    )
    ws_summary["A2"].font = SUBTITLE_FONT

    ws_summary["A4"] = "Selected Month"
    ws_summary["B4"] = months[0]
    ws_summary["B4"].number_format = "yyyy-mm"
    ws_summary["A4"].font = Font(bold=True)
    ws_summary["B4"].fill = VALUE_FILL

    total_month_count = max(len(months), 1)

    kpi_rows = [
        (
            "Total Base Cost (All Months)",
            f"=SUM('monthly cost breakdown'!$G$2:$G${monthly_last_data_row})",
        ),
        (
            "Total Buffered Cost (All Months)",
            f"=SUM('monthly cost breakdown'!$I$2:$I${monthly_last_data_row})",
        ),
        ("Average Monthly Buffered Cost", f"=IFERROR(B7/{total_month_count},0)"),
        (
            "Selected Month Buffered Cost",
            (
                "=SUMIFS('monthly cost breakdown'!$I$2:$I$"
                f"{monthly_last_data_row},'monthly cost breakdown'!$A$2:$A$"
                f"{monthly_last_data_row},$B$4)"
            ),
        ),
    ]

    start_kpi_row = 6
    for offset, (label, formula) in enumerate(kpi_rows):
        row = start_kpi_row + offset
        ws_summary.cell(row=row, column=1, value=label)
        value_cell = ws_summary.cell(row=row, column=2, value=formula)
        ws_summary.cell(row=row, column=1).font = Font(bold=True)
        value_cell.number_format = "$#,##0.00"
        value_cell.fill = VALUE_FILL

    service_header_row = 12
    ws_summary.cell(row=service_header_row, column=1, value="AWS Service")
    ws_summary.cell(row=service_header_row, column=2, value="Base Cost (USD)")
    ws_summary.cell(row=service_header_row, column=3, value="Buffered Cost (USD)")
    ws_summary.cell(row=service_header_row, column=4, value="Share of Buffered Total")

    for col in range(1, 5):
        cell = ws_summary.cell(row=service_header_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    service_start_row = service_header_row + 1

    for index, service in enumerate(services):
        row = service_start_row + index
        ws_summary.cell(row=row, column=1, value=service)
        ws_summary.cell(
            row=row,
            column=2,
            value=(
                "=SUMIFS('monthly cost breakdown'!$G$2:$G$"
                f"{monthly_last_data_row},'monthly cost breakdown'!$C$2:$C$"
                f"{monthly_last_data_row},A{row})"
            ),
        )
        ws_summary.cell(
            row=row,
            column=3,
            value=(
                "=SUMIFS('monthly cost breakdown'!$I$2:$I$"
                f"{monthly_last_data_row},'monthly cost breakdown'!$C$2:$C$"
                f"{monthly_last_data_row},A{row})"
            ),
        )
        ws_summary.cell(row=row, column=4, value=f"=IFERROR(C{row}/$B$7,0)")

        ws_summary.cell(row=row, column=2).number_format = "$#,##0.00"
        ws_summary.cell(row=row, column=3).number_format = "$#,##0.00"
        ws_summary.cell(row=row, column=4).number_format = "0.00%"

    trend_header_row = 12
    ws_summary.cell(row=trend_header_row, column=6, value="Month")
    ws_summary.cell(row=trend_header_row, column=7, value="Buffered Cost (USD)")

    for col in (6, 7):
        cell = ws_summary.cell(row=trend_header_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    trend_start_row = trend_header_row + 1
    for index, month_value in enumerate(months):
        row = trend_start_row + index
        ws_summary.cell(row=row, column=6, value=month_value)
        ws_summary.cell(
            row=row,
            column=7,
            value=(
                "=SUMIFS('monthly cost breakdown'!$I$2:$I$"
                f"{monthly_last_data_row},'monthly cost breakdown'!$A$2:$A$"
                f"{monthly_last_data_row},F{row})"
            ),
        )

        ws_summary.cell(row=row, column=6).number_format = "yyyy-mm"
        ws_summary.cell(row=row, column=7).number_format = "$#,##0.00"

    service_end_row = service_start_row + len(services) - 1
    trend_end_row = trend_start_row + len(months) - 1

    if len(services) > 0:
        pie_chart = PieChart()
        pie_data = Reference(ws_summary, min_col=3, min_row=service_header_row, max_row=service_end_row)
        pie_labels = Reference(ws_summary, min_col=1, min_row=service_start_row, max_row=service_end_row)
        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_labels)
        pie_chart.title = "Buffered Cost Mix by Service"
        pie_chart.height = 7
        pie_chart.width = 10
        ws_summary.add_chart(pie_chart, "E2")

    trend_chart = LineChart()
    trend_data = Reference(ws_summary, min_col=7, min_row=trend_header_row, max_row=trend_end_row)
    trend_categories = Reference(ws_summary, min_col=6, min_row=trend_start_row, max_row=trend_end_row)
    trend_chart.add_data(trend_data, titles_from_data=True)
    trend_chart.set_categories(trend_categories)
    trend_chart.title = "Buffered Cost Trend"
    trend_chart.y_axis.title = "USD"
    trend_chart.x_axis.title = "Month"
    trend_chart.height = 7
    trend_chart.width = 10
    ws_summary.add_chart(trend_chart, "E19")

    set_column_widths(
        ws_summary,
        {
            "A": 38,
            "B": 22,
            "C": 22,
            "D": 22,
            "E": 4,
            "F": 14,
            "G": 20,
            "H": 4,
            "I": 4,
        },
    )

    apply_borders(ws_summary, start_row=4, end_row=max(service_end_row, 10), start_col=1, end_col=4)
    apply_borders(ws_summary, start_row=trend_header_row, end_row=trend_end_row, start_col=6, end_col=7)


def generate_workbook(output_path: Path, terraform_dir: Path, months_count: int) -> Path:
    terraform_context = discover_terraform_context(terraform_dir)

    assumption_rows = build_assumption_rows(terraform_context)
    assumption_values = {row[0]: row[1] for row in assumption_rows}
    components = build_cost_components(terraform_context, assumption_values)

    wb = Workbook()
    ws_assumptions = wb.active
    ws_assumptions.title = "assumptions"
    ws_monthly = wb.create_sheet("monthly cost breakdown")
    ws_summary = wb.create_sheet("summary dashboard")
    ws_price = wb.create_sheet("aws price reference")

    start_month = date.today().replace(day=1)
    months = month_sequence(start_month=start_month, count=months_count)

    assumptions_row_lookup = build_assumptions_sheet(ws_assumptions, assumption_rows)
    build_price_reference_sheet(ws_price, terraform_context, components)
    services, monthly_last_data_row = build_monthly_cost_sheet(ws_monthly, months, components, assumptions_row_lookup)
    build_summary_dashboard(ws_summary, months, services, monthly_last_data_row, terraform_context)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return output_path


def parse_args() -> ArgumentParser:
    parser = ArgumentParser(description="Generate AWS cost estimation workbook using Terraform resources.")
    default_root = Path(__file__).resolve().parents[1]

    parser.add_argument(
        "--output",
        type=Path,
        default=default_root / "AWS_Cost_Estimation_Template.xlsx",
        help="Output .xlsx path",
    )
    parser.add_argument(
        "--terraform-dir",
        type=Path,
        default=default_root / "terraform",
        help="Terraform root folder used for discovery",
    )
    parser.add_argument(
        "--months",
        type=int,
        default=12,
        help="Number of monthly periods to generate",
    )

    return parser


def main() -> None:
    parser = parse_args()
    args = parser.parse_args()

    if args.months < 1:
        raise ValueError("--months must be at least 1")

    output_path = generate_workbook(args.output, args.terraform_dir, args.months)
    print(f"Workbook generated: {output_path}")


if __name__ == "__main__":
    main()
